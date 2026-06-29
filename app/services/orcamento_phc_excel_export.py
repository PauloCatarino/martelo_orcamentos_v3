"""Gerador do Excel do orçamento no FORMATO PHC (port do Martelo V2).

Reproduz a folha "PHC" do V2 (``_export_excel_phc_full``), confirmada pelo
modelo real ``260618_01_PHC.xlsx``, para ser importada pelo PHC.

Cada item gera uma linha principal (RefCliente/Referencia/Designacao + dimensões
numéricas + Qtd/Und/Venda) e, por cada linha extra da descrição, uma linha só
com a coluna ``Designacao`` (C) preenchida. A coluna ``Venda`` é escrita como
TEXTO ("1191,62", vírgula decimal) com ``number_format`` ``"@"`` para o PHC a ler
tal e qual.

Recebe DADOS simples (read-models ou ``SimpleNamespace``), sem DB nem Qt, para
ser testável.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from app.domain.descricao_format import parse_descricao

# Cabeçalho da folha "PHC" (colunas A..I), na ordem esperada pelo PHC.
_HEADERS = [
    "RefCliente",
    "Referencia",
    "Designacao",
    "XAltura",
    "YLargura",
    "ZEspessura",
    "Qtd",
    "Und",
    "Venda",
]
_PREFIXO = "COMP. MOB. - "
_REFERENCIA = "MOB"
_FMT_TEXTO = "@"


def _num(value) -> float | None:
    """Converte Decimal/número para float (None se vazio/inválido)."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _venda_texto(value) -> str | None:
    """Preço unitário -> texto "1191,62" (vírgula decimal, 2 casas).

    Devolve None quando o valor é vazio/inválido.
    """
    num = _num(value)
    if num is None:
        return None
    try:
        return f"{Decimal(str(num)):.2f}".replace(".", ",")
    except Exception:
        return None


def gerar_excel_phc(output_path, *, orcamento, items) -> Path:
    """Gera o Excel no formato PHC em ``output_path`` e devolve o ``Path``."""
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "PHC"
    ws.append(_HEADERS)

    for item in items:
        linhas = parse_descricao(getattr(item, "descricao", None))
        titulo = linhas[0].texto if (linhas and linhas[0].tipo != "vazia") else ""
        designacao = f"{_PREFIXO}{titulo.upper()}" if titulo else _PREFIXO.rstrip()

        und = (getattr(item, "unidade", "") or "").strip() or "un"
        und = "un" if und.lower() == "und" else und

        venda = _venda_texto(getattr(item, "preco_unitario", None))

        ws.append(
            [
                getattr(item, "codigo", None) or "",
                _REFERENCIA,
                designacao,
                _num(getattr(item, "altura", None)),
                _num(getattr(item, "largura", None)),
                _num(getattr(item, "profundidade", None)),
                _num(getattr(item, "quantidade", None)),
                und,
                venda,
            ]
        )
        if venda is not None:
            ws.cell(row=ws.max_row, column=9).number_format = _FMT_TEXTO

        # Linhas extra da descrição: só a coluna Designacao (C), restantes vazias.
        for linha in linhas[1:]:
            if linha.tipo == "vazia":
                continue
            if linha.tipo == "traco":
                texto = f"- {linha.texto}"
            elif linha.tipo == "estrela":
                texto = f"* {linha.texto}"
            elif linha.tipo == "titulo":
                texto = linha.texto.upper()
            else:
                texto = linha.texto
            ws.append(["", "", texto, None, None, None, None, None, None])

    # Ajuste simples da largura das colunas (como no V2).
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            max(max_len + 2, 10), 60
        )

    wb.save(str(output_path))

    return output_path
