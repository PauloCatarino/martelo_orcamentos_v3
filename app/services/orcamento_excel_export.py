"""Gerador do Excel do orçamento para o cliente (fase 8W.4.2).

Reproduz a folha "Relatório" do Martelo V2 (título "Orçamento"): bloco de
cabeçalho (título/nº/data + dados do cliente + Ref./Obra), tabela de items e
bloco de totais. Ao contrário do PDF, escreve NÚMEROS reais nas células
numéricas (com formato de número), não strings.

A função recebe DADOS simples (sem DB nem Qt) para ser testável.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.domain.descricao_format import parse_descricao
from app.utils.formatters import format_quantity, format_version

# Formatos de número das células (dimensões inteiras, Qt com casas opcionais,
# valores monetários em euros).
_FORMATO_DIM = "0"
_FORMATO_QT = "0.##"
_FORMATO_EUR = "#,##0.00 €"

# Cabeçalhos da tabela de items (colunas A..J).
_CABECALHOS = (
    "Item",
    "Código",
    "Descrição",
    "Alt",
    "Larg",
    "Prof",
    "Und",
    "Qt",
    "Preço Unit",
    "Preço Total",
)

# Larguras de coluna razoáveis (Descrição larga; restantes ajustadas).
_LARGURAS = {
    "A": 6,
    "B": 14,
    "C": 60,
    "D": 8,
    "E": 8,
    "F": 8,
    "G": 8,
    "H": 8,
    "I": 14,
    "J": 14,
}


def _format_data(value) -> str:
    """Formata a data do orçamento (datetime -> YYYY-MM-DD; senão str)."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _to_float(value) -> float | None:
    """Converte Decimal/número para float (None se vazio/inválido)."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _escrever_numero(cell, value, number_format: str) -> None:
    """Escreve um número (real) na célula com o formato dado."""
    numero = _to_float(value)
    if numero is not None:
        cell.value = numero
    cell.number_format = number_format


def _descricao_richtext(nome: str, descricao: str | None) -> CellRichText:
    """Descricao da coluna C: nome (negrito) + linhas formatadas."""
    blocos: list[tuple[InlineFont, str]] = [(InlineFont(b=True), nome)]
    for linha in parse_descricao(descricao):
        if linha.tipo == "titulo":
            blocos.append((InlineFont(b=True), linha.texto))
        elif linha.tipo == "traco":
            blocos.append((InlineFont(i=True), f"  - {linha.texto}"))
        elif linha.tipo == "estrela":
            blocos.append((InlineFont(i=True, color="FF0A5C0A"), f"  {linha.texto}"))
        else:
            blocos.append((InlineFont(), ""))

    runs = [
        TextBlock(fonte, ("" if i == 0 else "\n") + texto)
        for i, (fonte, texto) in enumerate(blocos)
    ]
    return CellRichText(runs)


def gerar_excel_orcamento(
    output_path,
    *,
    cliente,
    orcamento,
    items,
    totais,
    titulo: str = "Orçamento",
):
    """Gera o Excel do orçamento em ``output_path`` e devolve o ``Path``.

    ``cliente``/``orcamento``/``items``/``totais`` são dados simples (read-models
    ou ``SimpleNamespace``), sem DB nem Qt.
    """
    output_path = Path(output_path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = titulo

    negrito = Font(bold=True)
    borda_fina = Border(
        left=Side(style="thin", color="FF808080"),
        right=Side(style="thin", color="FF808080"),
        top=Side(style="thin", color="FF808080"),
        bottom=Side(style="thin", color="FF808080"),
    )
    cinza = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    # Bloco de cabeçalho (D1/D2/A3/D3/A4/A5/A6/A7/D7).
    num_versao_fmt = (
        f"{orcamento.num_orcamento}_{format_version(orcamento.numero_versao)}"
    )
    data_str = _format_data(getattr(orcamento, "created_at", None))

    ws["D1"] = titulo
    ws["D1"].font = Font(bold=True, size=14)
    ws["D2"] = f"Nº Orçamento: {num_versao_fmt}"
    ws["A3"] = cliente.nome or ""
    ws["A3"].font = negrito
    ws["D3"] = f"Data: {data_str}"
    ws["A4"] = getattr(cliente, "morada", None) or ""
    ws["A5"] = getattr(cliente, "email", None) or ""
    ws["A6"] = (
        f"Telefone: {getattr(cliente, 'telefone', None) or ''}  |  "
        f"N.º cliente PHC: {getattr(cliente, 'num_cliente', None) or ''}"
    )
    ws["A7"] = f"Ref.: {getattr(orcamento, 'ref_cliente', None) or '-'}"
    ws["D7"] = f"Obra: {getattr(orcamento, 'obra', None) or '-'}"

    # Linha 10: cabeçalhos a negrito com fundo cinza claro.
    for coluna, nome in enumerate(_CABECALHOS, start=1):
        cell = ws.cell(row=10, column=coluna, value=nome)
        cell.font = negrito
        cell.fill = cinza

    # Linha 11+: um item por linha (texto e números reais).
    linha = 11
    for item in items:
        ws.cell(row=linha, column=1, value=str(getattr(item, "ordem", "")))
        codigo_item = getattr(item, "codigo", None)
        nome_item = getattr(item, "item", None) or ""
        nome = f"{codigo_item} - {nome_item}" if codigo_item else nome_item
        ws.cell(row=linha, column=2, value=codigo_item or "")
        descricao_cell = ws.cell(row=linha, column=3)
        descricao_cell.value = _descricao_richtext(
            nome, getattr(item, "descricao", None)
        )
        descricao_cell.alignment = Alignment(wrap_text=True, vertical="top")
        _escrever_numero(ws.cell(row=linha, column=4), getattr(item, "altura", None), _FORMATO_DIM)
        _escrever_numero(ws.cell(row=linha, column=5), getattr(item, "largura", None), _FORMATO_DIM)
        _escrever_numero(ws.cell(row=linha, column=6), getattr(item, "profundidade", None), _FORMATO_DIM)
        ws.cell(row=linha, column=7, value=getattr(item, "unidade", None) or "")
        _escrever_numero(ws.cell(row=linha, column=8), getattr(item, "quantidade", None), _FORMATO_QT)
        _escrever_numero(ws.cell(row=linha, column=9), getattr(item, "preco_unitario", None), _FORMATO_EUR)
        _escrever_numero(ws.cell(row=linha, column=10), getattr(item, "preco_total", None), _FORMATO_EUR)
        ws.cell(row=linha, column=10).font = negrito
        linha += 1

    ultima_linha_items = max(10, linha - 1)
    for linha_tabela in range(10, ultima_linha_items + 1):
        for coluna in range(1, 11):
            ws.cell(row=linha_tabela, column=coluna).border = borda_fina

    for linha_item in range(11, ultima_linha_items + 1):
        ws.cell(row=linha_item, column=1).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        ws.cell(row=linha_item, column=7).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        for coluna in (4, 5, 6, 8, 9, 10):
            ws.cell(row=linha_item, column=coluna).alignment = Alignment(
                horizontal="right", vertical="top"
            )

    # Totais nas colunas I/J, deixando 1 linha em branco após os items.
    iva_label = f"IVA ({format_quantity(totais.iva_pct)}%):"
    totais_linhas = (
        ("Total Qt:", totais.total_qt, _FORMATO_QT),
        ("SubTotal:", totais.subtotal, _FORMATO_EUR),
        (iva_label, totais.iva, _FORMATO_EUR),
        ("Total Geral:", totais.total_geral, _FORMATO_EUR),
    )
    primeira_linha_totais = linha + 1
    for deslocamento, (rotulo, valor, formato) in enumerate(totais_linhas):
        linha_total = primeira_linha_totais + deslocamento
        rotulo_cell = ws.cell(row=linha_total, column=9, value=rotulo)
        rotulo_cell.font = negrito
        _escrever_numero(ws.cell(row=linha_total, column=10), valor, formato)

    for coluna, largura in _LARGURAS.items():
        ws.column_dimensions[coluna].width = largura

    workbook.save(str(output_path))

    return output_path
