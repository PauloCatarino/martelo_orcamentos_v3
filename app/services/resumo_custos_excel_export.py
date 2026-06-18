"""Gerador do Excel interno "Resumo de Custos" a partir do modelo (8W.4.3/4.4).

Copia um ficheiro modelo ``.xlsx`` ja formatado e preenche apenas as folhas de
resumo. A funcao recebe dados simples (sem DB nem Qt) para ser testavel.
"""

from __future__ import annotations

import shutil
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

_FMT_DIM = "0"
_FMT_INT = "0"
_FMT_QT = "0.##"
_FMT_M2 = "0.000"
_FMT_ML = "0.00"
_FMT_EUR = "#,##0.00 €"
_FMT_PCT = "0.0"

COLUNAS_RESUMO_GERAL = [
    "id",
    "item_id",
    "tipo_linha",
    "def_peca",
    "descricao",
    "descricao_livre",
    "mat_default",
    "quantidade",
    "item_qt",
    "qt_total",
    "comp",
    "larg",
    "esp",
    "comp_real",
    "larg_real",
    "esp_real",
    "area_m2_und",
    "perimetro_ml_und",
    "ref_le",
    "descricao_no_orcamento",
    "unidade",
    "pliq",
    "desp",
    "tipo_mp",
    "familia_mp",
    "comp_mp",
    "larg_mp",
    "esp_mp",
    "coresp_orla_0_4",
    "coresp_orla_1_0",
    "ml_orla_fina",
    "ml_orla_grossa",
    "custo_orla_fina",
    "custo_orla_grossa",
    "custo_orlas",
    "consumo_ml_total",
    "acabamento_sup",
    "acabamento_inf",
    "custo_acabamento",
    "operacoes",
    "custo_mp",
    "custo_ferragem",
    "custo_corte",
    "custo_orlagem",
    "custo_cnc",
    "custo_montagem_manual",
    "custo_producao",
    "custo_total",
    "excluir_mp",
    "excluir_orla",
    "excluir_ferragem",
    "excluir_producao",
    "excluir_acabamento",
    "excluir_mo",
]

_CAMPOS_MULTIPLICAR = {
    "ml_orla_fina",
    "ml_orla_grossa",
    "custo_orla_fina",
    "custo_orla_grossa",
    "custo_orlas",
    "consumo_ml_total",
    "custo_acabamento",
    "custo_mp",
    "custo_ferragem",
    "custo_corte",
    "custo_orlagem",
    "custo_cnc",
    "custo_montagem_manual",
    "custo_producao",
    "custo_total",
}

_CAMPOS_BOOL_GERAL = {
    "excluir_mp",
    "excluir_orla",
    "excluir_ferragem",
    "excluir_producao",
    "excluir_acabamento",
    "excluir_mo",
}

_MAPA_ATRIBUTOS_GERAL = {
    "item_id": "orcamento_item_id",
    "def_peca": "def_peca_codigo",
    "tipo_mp": "tipo_materia_prima",
    "familia_mp": "familia_materia_prima",
    "pliq": "preco_liquido",
    "desp": "desperdicio_percentagem",
    "area_m2_und": "area_m2",
    "perimetro_ml_und": "perimetro_ml",
    "acabamento_sup": "acabamento_face_sup",
    "acabamento_inf": "acabamento_face_inf",
}

_CAMPOS_INT_GERAL = {"id", "item_id", "qt_total"} | _CAMPOS_BOOL_GERAL
_CAMPOS_QT_GERAL = {"quantidade", "item_qt"}
_CAMPOS_DIM_GERAL = {
    "comp_real",
    "larg_real",
    "esp_real",
    "comp_mp",
    "larg_mp",
    "esp_mp",
}
_CAMPOS_M2_GERAL = {"area_m2_und"}
_CAMPOS_ML_GERAL = {"perimetro_ml_und", "ml_orla_fina", "ml_orla_grossa", "consumo_ml_total"}
_CAMPOS_PCT_GERAL = {"desp"}

_ZERO = Decimal("0")
_UM = Decimal("1")


def _to_float(value) -> float | None:
    """Converte Decimal/numero para float (None se vazio/invalido)."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_decimal(value) -> Decimal | None:
    """Converte valores numericos para Decimal, mantendo None se invalido."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _multiplicar(value, fator: Decimal):
    numero = _to_decimal(value)
    if numero is None:
        return None
    return numero * fator


def _escrever_numero(cell, value, number_format: str) -> None:
    """Escreve um numero real na celula com o formato dado."""
    numero = _to_float(value)
    if numero is not None:
        cell.value = numero
    cell.number_format = number_format


def _texto(value) -> str:
    return "" if value is None else str(value)


def _limpar_linhas_dados(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _folha(wb, nome: str):
    if nome not in wb.sheetnames:
        return None
    ws = wb[nome]
    _limpar_linhas_dados(ws)
    return ws


def construir_linhas_geral(linhas, item_qt_por_item):
    """Devolve dicts para o dump "Resumo Geral" a partir das linhas ativas."""
    linhas_geral = []
    item_qt_por_item = item_qt_por_item or {}

    for linha in linhas or []:
        if not getattr(linha, "ativo", True):
            continue

        item_id = getattr(linha, "orcamento_item_id", None)
        item_qt = _to_decimal(item_qt_por_item.get(item_id)) or _UM
        quantidade = _to_decimal(getattr(linha, "quantidade", None)) or _ZERO

        valores = {}
        for coluna in COLUNAS_RESUMO_GERAL:
            if coluna == "item_id":
                valor = item_id
            elif coluna == "item_qt":
                valor = item_qt
            elif coluna == "qt_total":
                valor = quantidade * item_qt
            elif coluna in _CAMPOS_BOOL_GERAL:
                valor = 1 if getattr(linha, coluna, False) else 0
            else:
                atributo = _MAPA_ATRIBUTOS_GERAL.get(coluna, coluna)
                valor = getattr(linha, atributo, None)
                if coluna in _CAMPOS_MULTIPLICAR:
                    valor = _multiplicar(valor, item_qt)

            valores[coluna] = valor

        linhas_geral.append(valores)

    return linhas_geral


def _formato_resumo_geral(coluna: str) -> str | None:
    if coluna in _CAMPOS_INT_GERAL:
        return _FMT_INT
    if coluna in _CAMPOS_QT_GERAL:
        return _FMT_QT
    if coluna in _CAMPOS_DIM_GERAL:
        return _FMT_DIM
    if coluna in _CAMPOS_M2_GERAL:
        return _FMT_M2
    if coluna in _CAMPOS_ML_GERAL:
        return _FMT_ML
    if coluna == "pliq" or coluna.startswith("custo_"):
        return _FMT_EUR
    if coluna in _CAMPOS_PCT_GERAL:
        return _FMT_PCT
    return None


def _preencher_resumo_geral(wb, linhas_geral) -> None:
    if linhas_geral is None or "Resumo Geral" not in wb.sheetnames:
        return

    ws = wb["Resumo Geral"]
    for col_idx, nome in enumerate(COLUNAS_RESUMO_GERAL, start=1):
        ws.cell(row=1, column=col_idx, value=nome)

    if ws.max_column > len(COLUNAS_RESUMO_GERAL):
        ws.delete_cols(
            len(COLUNAS_RESUMO_GERAL) + 1,
            ws.max_column - len(COLUNAS_RESUMO_GERAL),
        )

    _limpar_linhas_dados(ws)

    for linha_idx, linha in enumerate(linhas_geral, start=2):
        for col_idx, coluna in enumerate(COLUNAS_RESUMO_GERAL, start=1):
            cell = ws.cell(row=linha_idx, column=col_idx)
            valor = linha.get(coluna)
            formato = _formato_resumo_geral(coluna)
            if formato is None:
                cell.value = _texto(valor)
            else:
                _escrever_numero(cell, valor, formato)


def _preencher_placas(wb, placas) -> None:
    ws = _folha(wb, "Resumo Placas")
    if ws is None:
        return

    for linha, placa in enumerate(placas or [], start=2):
        ws.cell(row=linha, column=1, value=_texto(getattr(placa, "ref_le", None)))
        ws.cell(
            row=linha,
            column=2,
            value=_texto(getattr(placa, "descricao_no_orcamento", None)),
        )
        _escrever_numero(ws.cell(row=linha, column=3), getattr(placa, "pliq", None), _FMT_EUR)
        ws.cell(row=linha, column=4, value=_texto(getattr(placa, "unidade", None)))
        _escrever_numero(ws.cell(row=linha, column=5), getattr(placa, "desp", None), _FMT_PCT)
        _escrever_numero(ws.cell(row=linha, column=6), getattr(placa, "comp_mp", None), _FMT_DIM)
        _escrever_numero(ws.cell(row=linha, column=7), getattr(placa, "larg_mp", None), _FMT_DIM)
        _escrever_numero(ws.cell(row=linha, column=8), getattr(placa, "esp_mp", None), _FMT_DIM)
        _escrever_numero(ws.cell(row=linha, column=9), getattr(placa, "qt_placas", None), _FMT_INT)
        _escrever_numero(ws.cell(row=linha, column=10), getattr(placa, "area_placa", None), _FMT_M2)
        _escrever_numero(ws.cell(row=linha, column=11), getattr(placa, "m2_consumidos", None), _FMT_M2)
        _escrever_numero(ws.cell(row=linha, column=12), getattr(placa, "custo_mp_total", None), _FMT_EUR)
        _escrever_numero(
            ws.cell(row=linha, column=13),
            getattr(placa, "custo_placa_inteira", None),
            _FMT_EUR,
        )
        _escrever_numero(
            ws.cell(row=linha, column=14),
            1 if getattr(placa, "nao_stock", False) else 0,
            _FMT_INT,
        )


def _preencher_orlas(wb, orlas) -> None:
    ws = _folha(wb, "Resumo Orlas")
    if ws is None:
        return

    for linha, orla in enumerate(orlas or [], start=2):
        ws.cell(row=linha, column=1, value=_texto(getattr(orla, "ref_orla", None)))
        _escrever_numero(ws.cell(row=linha, column=2), getattr(orla, "espessura", None), _FMT_QT)
        _escrever_numero(ws.cell(row=linha, column=3), getattr(orla, "largura", None), _FMT_DIM)
        _escrever_numero(ws.cell(row=linha, column=4), getattr(orla, "ml_total", None), _FMT_ML)
        _escrever_numero(ws.cell(row=linha, column=5), getattr(orla, "custo_total", None), _FMT_EUR)


def _custo_unitario_ferragem(ferragem):
    qt_total = getattr(ferragem, "qt_total", None)
    if not qt_total:
        return 0
    return getattr(ferragem, "custo_total", 0) / qt_total


def _preencher_ferragens(wb, ferragens) -> None:
    ws = _folha(wb, "Resumo Ferragens")
    if ws is None:
        return

    for linha, ferragem in enumerate(ferragens or [], start=2):
        ws.cell(row=linha, column=1, value=_texto(getattr(ferragem, "ref_le", None)))
        ws.cell(
            row=linha,
            column=2,
            value=_texto(getattr(ferragem, "descricao_no_orcamento", None)),
        )
        _escrever_numero(ws.cell(row=linha, column=3), getattr(ferragem, "pliq", None), _FMT_EUR)
        ws.cell(row=linha, column=4, value=_texto(getattr(ferragem, "unidade", None)))
        _escrever_numero(ws.cell(row=linha, column=5), getattr(ferragem, "desp", None), _FMT_PCT)
        ws.cell(row=linha, column=6, value="")
        ws.cell(row=linha, column=7, value="")
        ws.cell(row=linha, column=8, value="")
        _escrever_numero(ws.cell(row=linha, column=9), getattr(ferragem, "qt_total", None), _FMT_QT)
        _escrever_numero(ws.cell(row=linha, column=10), getattr(ferragem, "ml", None), _FMT_ML)
        _escrever_numero(ws.cell(row=linha, column=11), _custo_unitario_ferragem(ferragem), _FMT_EUR)
        _escrever_numero(ws.cell(row=linha, column=12), getattr(ferragem, "custo_total", None), _FMT_EUR)


def _preencher_maquinas(wb, maquinas) -> None:
    ws = _folha(wb, "Resumo Maquinas_MO")
    if ws is None:
        return

    for linha, maquina in enumerate(maquinas or [], start=2):
        ws.cell(row=linha, column=1, value=_texto(getattr(maquina, "centro", None)))
        _escrever_numero(ws.cell(row=linha, column=2), getattr(maquina, "custo_total", None), _FMT_EUR)
        _escrever_numero(ws.cell(row=linha, column=3), getattr(maquina, "ml_corte", None), _FMT_ML)
        _escrever_numero(ws.cell(row=linha, column=4), getattr(maquina, "ml_orlado", None), _FMT_ML)
        _escrever_numero(ws.cell(row=linha, column=5), getattr(maquina, "num_pecas", None), _FMT_QT)


def _preencher_margens(wb, distribuicao) -> None:
    ws = _folha(wb, "Resumo Margens")
    if ws is None:
        return

    linha = 2
    for categoria in getattr(distribuicao, "categorias", None) or []:
        ws.cell(row=linha, column=1, value=_texto(getattr(categoria, "nome", None)))
        _escrever_numero(ws.cell(row=linha, column=2), getattr(categoria, "pct", None), _FMT_PCT)
        _escrever_numero(ws.cell(row=linha, column=3), getattr(categoria, "euros", None), _FMT_EUR)
        linha += 1

    ws.cell(row=linha, column=1, value="Total (Venda)")
    ws.cell(row=linha, column=2, value="")
    _escrever_numero(
        ws.cell(row=linha, column=3),
        getattr(distribuicao, "total_venda", None),
        _FMT_EUR,
    )


def gerar_excel_resumo_custos(output_path, template_path, *, resumo, linhas_geral=None):
    """Copia o MODELO para ``output_path`` e preenche as folhas de resumo.

    Preserva o ficheiro modelo (incluindo cabecalhos e formatacao existentes),
    escreve dados a partir da linha 2 e devolve o ``Path`` de saida.
    """
    output_path = Path(output_path)
    template_path = Path(template_path)

    shutil.copyfile(template_path, output_path)
    workbook = load_workbook(output_path)

    _preencher_resumo_geral(workbook, linhas_geral)
    _preencher_placas(workbook, getattr(resumo, "placas", None))
    _preencher_orlas(workbook, getattr(resumo, "orlas", None))
    _preencher_ferragens(workbook, getattr(resumo, "ferragens", None))
    _preencher_maquinas(workbook, getattr(resumo, "maquinas", None))
    _preencher_margens(workbook, getattr(resumo, "distribuicao", None))

    workbook.save(str(output_path))
    return output_path
