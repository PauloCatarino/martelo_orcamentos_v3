"""Indexador RAG dos catalogos de fornecedores para a Pesquisa IA."""

from __future__ import annotations

import json
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.system_setting_service import SystemSettingService

EMBEDDINGS_FILENAME = "embeddings.npy"
META_FILENAME = "meta.jsonl"
EXTENSOES = (".xlsx", ".xlsm", ".pdf")
MODELO_EMBEDDINGS_DEFAULT = "paraphrase-multilingual-MiniLM-L12-v2"


@dataclass(frozen=True)
class ResultadoIndexacao:
    ficheiros: int
    chunks: int
    erros: int
    pasta_indice: str


def _config(session: Session) -> tuple[str, str, str]:
    svc = SystemSettingService(session)
    catalogos = (svc.obter_valor("pasta_pesquisa_profunda_ia", "") or "").strip()
    indice = (svc.obter_valor("pasta_embeddings_ia", "") or "").strip()
    modelo = (
        (svc.obter_valor("modelo_embeddings_ia", "") or "").strip()
        or MODELO_EMBEDDINGS_DEFAULT
    )
    return catalogos, indice, modelo


def _chunks_excel(caminho: Path) -> Iterator[tuple[str, dict]]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        for folha in wb.sheetnames:
            ws = wb[folha]
            cabecalho: list[str] | None = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                valores = [
                    "" if celula is None else str(celula).strip() for celula in row
                ]
                nao_vazias = [valor for valor in valores if valor]
                if not nao_vazias:
                    continue
                if cabecalho is None:
                    if len(nao_vazias) >= 4:
                        cabecalho = valores
                    continue
                partes = [
                    f"{coluna}: {valor}" if coluna else valor
                    for coluna, valor in zip(cabecalho, valores)
                    if valor
                ]
                if partes:
                    yield " | ".join(partes), {"folha": folha, "linha": i}
    finally:
        wb.close()


def _chunks_pdf(caminho: Path) -> Iterator[tuple[str, dict]]:
    from pypdf import PdfReader

    reader = PdfReader(str(caminho))
    for i, page in enumerate(reader.pages, start=1):
        texto = (page.extract_text() or "").strip()
        if texto:
            yield texto, {"pagina": i}


def indexar(
    session: Session, progresso: Callable[[str], None] | None = None
) -> ResultadoIndexacao:
    catalogos, indice, modelo_nome = _config(session)
    if not catalogos:
        raise RuntimeError("A 'Pasta Pesquisa Profunda IA' nao esta configurada.")
    base = Path(catalogos)
    if not base.exists():
        raise RuntimeError(f"Pasta de catalogos nao acessivel: {catalogos}")
    if not indice:
        raise RuntimeError("A 'Pasta Embeddings IA' nao esta configurada.")
    destino = Path(indice)
    destino.mkdir(parents=True, exist_ok=True)

    try:
        import numpy as np
        from sentence_transformers import SentenceTransformer
    except ImportError as exc:
        raise RuntimeError(
            "Faltam dependencias de IA. Instale: pip install sentence-transformers pypdf"
        ) from exc

    textos: list[str] = []
    metadados: list[dict] = []
    ficheiros = 0
    erros = 0
    for caminho in sorted(base.rglob("*")):
        if not caminho.is_file() or caminho.suffix.lower() not in EXTENSOES:
            continue
        fornecedor = caminho.parent.name
        try:
            gerador = (
                _chunks_excel(caminho)
                if caminho.suffix.lower() in (".xlsx", ".xlsm")
                else _chunks_pdf(caminho)
            )
            for texto, extra in gerador:
                textos.append(texto)
                metadados.append(
                    {
                        "ficheiro": caminho.name,
                        "caminho": str(caminho),
                        "fornecedor": fornecedor,
                        "texto": texto[:600],
                        **extra,
                    }
                )
            ficheiros += 1
            if progresso:
                progresso(f"{caminho.name}: {len(textos)} chunks acumulados")
        except Exception:  # noqa: BLE001
            erros += 1

    if not textos:
        raise RuntimeError("Nenhum conteudo extraido dos catalogos (Excel/PDF).")

    if progresso:
        progresso(f"A gerar embeddings de {len(textos)} chunks (modelo {modelo_nome})...")
    modelo = SentenceTransformer(modelo_nome)
    vetores = modelo.encode(
        textos, normalize_embeddings=True, show_progress_bar=False, batch_size=64
    ).astype("float32")

    np.save(destino / EMBEDDINGS_FILENAME, vetores)
    with open(destino / META_FILENAME, "w", encoding="utf-8") as ficheiro_meta:
        for meta in metadados:
            ficheiro_meta.write(json.dumps(meta, ensure_ascii=False) + "\n")

    return ResultadoIndexacao(
        ficheiros=ficheiros, chunks=len(textos), erros=erros, pasta_indice=str(destino)
    )
