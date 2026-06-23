"""Read the curated board references Excel file for Pesquisa IA."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.system_setting_service import SystemSettingService
from app.utils.formatters import format_currency

FICHEIRO_REFERENCIAS = "12_Placas_Referencias_COMPLETO.xlsx"


@dataclass(frozen=True)
class LinhaReferencia:
    folha: str
    referencia: str
    st_acab: str
    nome_design: str
    grupo: str
    tipo: str
    fornecedor: str
    precos: dict[str, str]


def _norm(value: object) -> str:
    if value is None:
        return ""
    texto = unicodedata.normalize("NFKD", str(value))
    texto = "".join(
        caractere for caractere in texto if not unicodedata.combining(caractere)
    )
    return texto.lower().strip()


def _idx(cabecalho: list[object], *subs: str) -> int | None:
    for index, header in enumerate(cabecalho):
        if any(sub in _norm(header) for sub in subs):
            return index
    return None


def _idx_st(cabecalho: list[object]) -> int | None:
    for index, header in enumerate(cabecalho):
        header_norm = _norm(header)
        if header_norm == "st" or "acab" in header_norm or "substrato" in header_norm:
            return index
    return None


def _val(valores: list[object], index: int | None) -> str:
    if index is None or index >= len(valores):
        return ""
    valor = valores[index]
    return "" if valor is None else str(valor).strip()


def _precos(cabecalho: list[object], valores: list[object]) -> dict[str, str]:
    out: dict[str, str] = {}
    for header, valor in zip(cabecalho, valores):
        if valor is None or str(valor).strip() == "":
            continue
        header_norm = _norm(header)
        if "preco tabela" not in header_norm and "pvp" not in header_norm:
            continue

        match = re.search(r"(\d+)\s*mm", str(header), re.IGNORECASE)
        etiqueta = f"{match.group(1)}mm" if match else str(header).strip()
        preco = format_currency(valor)
        if preco:
            out[etiqueta] = preco
    return out


def listar_referencias(session: Session) -> list[LinhaReferencia]:
    base = (
        SystemSettingService(session).obter_valor("pasta_pesquisa_profunda_ia", "")
        or ""
    ).strip()
    caminho = Path(base) / FICHEIRO_REFERENCIAS
    if not caminho.exists():
        raise RuntimeError(f"Excel de referencias nao encontrado: {caminho}")

    workbook = load_workbook(caminho, read_only=True, data_only=True)
    linhas: list[LinhaReferencia] = []
    try:
        for folha in workbook.sheetnames:
            worksheet = workbook[folha]
            cabecalho: list[object] | None = None
            for row in worksheet.iter_rows(values_only=True):
                valores = ["" if celula is None else str(celula).strip() for celula in row]
                nao_vazias = [valor for valor in valores if valor]
                if not nao_vazias:
                    continue

                if cabecalho is None:
                    if len(nao_vazias) >= 4 and any(
                        "refer" in _norm(valor) for valor in valores
                    ):
                        cabecalho = list(row)
                    continue

                referencia = _val(valores, _idx(cabecalho, "refer"))
                if not referencia:
                    continue

                linhas.append(
                    LinhaReferencia(
                        folha=folha,
                        referencia=referencia,
                        st_acab=_val(valores, _idx_st(cabecalho)),
                        nome_design=_val(valores, _idx(cabecalho, "nome", "descri")),
                        grupo=_val(valores, _idx(cabecalho, "grupo")),
                        tipo=_val(
                            valores,
                            _idx(cabecalho, "tipo produto", "familia produto"),
                        ),
                        fornecedor=_val(valores, _idx(cabecalho, "fornec")),
                        precos=_precos(cabecalho, valores),
                    )
                )
    finally:
        workbook.close()

    return linhas
