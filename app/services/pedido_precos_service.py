"""Prepare price-update requests to suppliers: the attachment and the email.

The email is **opened in Outlook, not sent**. These messages go to people
outside the company, and the address kept on the supplier is only a suggestion —
the user reads the message, fixes the recipient, adds whoever else should be in
copy, and sends it themselves.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

from sqlalchemy.orm import Session

from app.domain.pedido_precos import (
    COLUNAS_A_PREENCHER,
    COLUNAS_ANEXO,
    PedidoFornecedor,
    agrupar_por_fornecedor,
    assunto_do_email,
    corpo_do_email,
    nome_do_anexo,
)
from app.services.def_fornecedor_service import DefFornecedorService
from app.services.def_materia_prima_service import DefMateriaPrimaService
from app.services.system_setting_service import SystemSettingService

SYSTEM_SETTING_PASTA_MATERIAS_PRIMAS = "pasta_materias_primas"
SUBPASTA_PEDIDOS = "Pedidos_Precos"

# Cores do anexo: castanho no cabeçalho, azul nas colunas a preencher.
COR_CABECALHO = "8B6F4E"
COR_A_PREENCHER = "D8E7F3"
COR_BLOQUEADA = "EFE7DA"


@dataclass(frozen=True)
class PedidoPreparado:
    """One request already written to disk and ready to be mailed."""

    pedido: PedidoFornecedor
    anexo: Path
    assunto: str
    corpo_html: str


class PedidoPrecosService:
    """Gather what needs reviewing, build the attachments and open the emails."""

    def __init__(self, session: Session) -> None:
        self.session = session

    def levantar_pedidos(
        self, meses_limite: int = 12, hoje: date | None = None
    ) -> list[PedidoFornecedor]:
        """Os pedidos por fornecedor, a partir do que está por rever."""
        materias = DefMateriaPrimaService(self.session).listar_materias_primas()
        fornecedores = DefFornecedorService(self.session).listar_fornecedores()

        return agrupar_por_fornecedor(materias, fornecedores, hoje, meses_limite)

    def pasta_dos_pedidos(self) -> Path:
        """Onde ficam guardados os anexos enviados.

        Ao lado do Excel das matérias-primas, para haver rasto do que foi pedido
        e quando. Sem essa pasta configurada, fica na pasta de trabalho.
        """
        configurada = SystemSettingService(self.session).obter_valor(
            SYSTEM_SETTING_PASTA_MATERIAS_PRIMAS, default=None
        )
        base = Path((configurada or "").strip()) if configurada else Path.cwd()

        return base / SUBPASTA_PEDIDOS

    def preparar(
        self,
        pedido: PedidoFornecedor,
        remetente: str | None = None,
        hoje: date | None = None,
        pasta: Path | None = None,
    ) -> PedidoPreparado:
        """Escrever o anexo e montar o texto do email (sem enviar nada)."""
        destino = pasta or self.pasta_dos_pedidos()
        destino.mkdir(parents=True, exist_ok=True)
        caminho = destino / nome_do_anexo(pedido, hoje)
        escrever_anexo(pedido, caminho)

        return PedidoPreparado(
            pedido=pedido,
            anexo=caminho,
            assunto=assunto_do_email(pedido),
            corpo_html=corpo_do_email(pedido, caminho.name, remetente),
        )

    def abrir_no_outlook(self, preparado: PedidoPreparado) -> None:
        """Abrir o email no Outlook, para o utilizador rever e enviar."""
        abrir_email_no_outlook(
            destino=preparado.pedido.email or "",
            assunto=preparado.assunto,
            corpo_html=preparado.corpo_html,
            anexos=[str(preparado.anexo)],
            cc=preparado.pedido.email_cc or "",
        )


def escrever_anexo(pedido: PedidoFornecedor, caminho: Path) -> Path:
    """Escrever o ficheiro que segue para o fornecedor.

    As colunas que ele preenche vão a azul e desbloqueadas; as restantes vão
    bloqueadas, para não serem alteradas por engano — em especial o «Código»,
    que é o que permite ao V3 saber a que material pertence cada resposta.
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill, Protection

    livro = openpyxl.Workbook()
    folha = livro.active
    folha.title = "Precos"

    for indice, coluna in enumerate(COLUNAS_ANEXO, start=1):
        celula = folha.cell(row=1, column=indice, value=coluna)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for numero, linha in enumerate(pedido.linhas, start=2):
        valores = linha.valores_do_anexo()
        for indice, coluna in enumerate(COLUNAS_ANEXO, start=1):
            valor = valores[coluna]
            celula = folha.cell(row=numero, column=indice)
            celula.value = float(valor) if hasattr(valor, "quantize") else valor
            a_preencher = coluna in COLUNAS_A_PREENCHER
            celula.fill = PatternFill(
                "solid", fgColor=COR_A_PREENCHER if a_preencher else COR_BLOQUEADA
            )
            celula.protection = Protection(locked=not a_preencher)

    larguras = {
        "A": 12, "B": 20, "C": 52, "D": 8, "E": 16,
        "F": 18, "G": 12, "H": 18, "I": 34, "J": 30,
    }
    for coluna, largura in larguras.items():
        folha.column_dimensions[coluna].width = largura

    folha.freeze_panes = "A2"
    # A folha vai protegida com as colunas a preencher livres: assim o
    # fornecedor não desalinha o ficheiro nem mexe no código do material.
    folha.protection.sheet = True
    folha.protection.password = "LE"
    folha.protection.enable()

    caminho.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho)

    return caminho


def abrir_email_no_outlook(
    destino: str,
    assunto: str,
    corpo_html: str,
    anexos: list[str] | None = None,
    cc: str = "",
) -> None:
    """Abrir uma mensagem no Outlook, já preenchida, sem a enviar.

    Ao contrário do envio dos orçamentos, aqui a mensagem fica aberta à espera
    de quem a escreve: o destinatário guardado é uma sugestão e estes emails vão
    para fora da empresa.
    """
    import os

    from app.services.email_service import _ligar_outlook, _require_win32com_client

    win32_client = _require_win32com_client()
    try:
        import pythoncom
    except Exception as exc:  # pragma: no cover - depende do pywin32 instalado
        raise RuntimeError(
            "Preparar o email requer o modulo 'pythoncom' do pacote pywin32."
        ) from exc

    pythoncom.CoInitialize()
    try:
        outlook = _ligar_outlook(win32_client)
        mail = outlook.CreateItem(0)
        mail.To = destino
        if cc:
            mail.CC = cc
        mail.Subject = assunto
        mail.HTMLBody = corpo_html
        for caminho in anexos or []:
            if os.path.exists(caminho):
                mail.Attachments.Add(caminho)
        mail.Display()
    finally:
        pythoncom.CoUninitialize()


def _require_openpyxl():
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - openpyxl faz parte do projeto
        raise RuntimeError(
            "openpyxl nao esta instalado. Instale com: pip install openpyxl"
        ) from error

    return openpyxl
