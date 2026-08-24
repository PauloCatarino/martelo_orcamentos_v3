"""Registo e exportação assistida dos PDFs da Lista Material.

O inventário é explícito: documentos sem folha/macro conhecida aparecem como
indisponíveis, em vez de se adivinharem intervalos de impressão.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.models.lista_material_assistente import (
    ListaMaterialPdfDocumento,
    ListaMaterialPdfPreset,
)


@dataclass(frozen=True)
class PdfDocument:
    identifier: str
    name: str
    category: str
    sheets: tuple[str, ...]
    filename: str
    order: int
    combinable: bool = True
    unavailable_reason: str = ""


@dataclass(frozen=True)
class PdfDocumentState:
    document: PdfDocument
    available: bool
    reason: str = ""
    export_sheets: tuple[str, ...] = ()


@dataclass(frozen=True)
class PdfExportResult:
    files: tuple[Path, ...]
    package: Path | None
    errors: tuple[str, ...]


DEFAULT_DOCUMENTS = (
    PdfDocument(
        "ferragens",
        "Ferragens + Purch + SPP",
        "Ferragens",
        ("1_FERRAGENS", "2_PURCH", "3_SPP"),
        "2_Lista_Ferragens_{nome_enc_imos}.pdf",
        20,
    ),
    PdfDocument(
        "resumo_orlas",
        "Resumo de Orlas",
        "Orlas",
        ("ResumoOrlas",),
        "4_Resumo_Orlas_{nome_enc_imos}.pdf",
        40,
    ),
    PdfDocument(
        "etiqueta_palete",
        "Etiqueta Palete",
        "Etiquetas e paletes",
        ("5_ETIQUETA_PALETE",),
        "5_Etiqueta_Palete_{nome_enc_imos}.pdf",
        50,
    ),
    PdfDocument(
        "listagem_artigo",
        "Listagem por Artigo",
        "Listagens",
        ("LISTAGEM_por_Artigo",),
        "6_Lista_Material_{nome_enc_imos}.pdf",
        60,
    ),
    PdfDocument(
        "relatorio",
        "Relatório geral",
        "Relatórios gerais",
        ("RELATORIO",),
        "Relatorio_Geral.pdf",
        80,
    ),
)

RETIRED_DOCUMENT_IDS = frozenset(
    {"caderno_encargos", "rosto", "purch", "spp", "listagem_cutrite"}
)
FERRAGENS_LEGACY_IDS = frozenset({"ferragens", "purch", "spp"})


def sync_pdf_document_registry(session: Session) -> int:
    """Regista o inventário conhecido sem remover definições personalizadas."""
    changed = 0
    for document in DEFAULT_DOCUMENTS:
        row = session.execute(
            select(ListaMaterialPdfDocumento).where(
                ListaMaterialPdfDocumento.identificador == document.identifier
            )
        ).scalar_one_or_none()
        if row is None:
            row = ListaMaterialPdfDocumento(identificador=document.identifier)
            session.add(row)
            changed += 1
        row.nome = document.name
        row.categoria = document.category
        row.origem_tipo = "folhas" if len(document.sheets) > 1 else "folha"
        row.origem_valor = "|".join(document.sheets)
        row.nome_ficheiro = document.filename
        row.combinavel = document.combinable
        row.ordem = document.order
        row.ativo = True
        row.prerequisitos_json = json.dumps(
            {"indisponivel": document.unavailable_reason}, ensure_ascii=False
        )
    session.execute(
        update(ListaMaterialPdfDocumento)
        .where(ListaMaterialPdfDocumento.identificador.in_(RETIRED_DOCUMENT_IDS))
        .values(ativo=False)
    )
    session.commit()
    return changed


def inspect_pdf_documents(workbook_path: Path) -> list[PdfDocumentState]:
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        sheets = set(workbook.sheetnames)
        result: list[PdfDocumentState] = []
        for document in sorted(DEFAULT_DOCUMENTS, key=lambda item: item.order):
            if document.unavailable_reason:
                result.append(PdfDocumentState(document, False, document.unavailable_reason))
                continue

            export_sheets: list[str] = []
            unavailable_sheets: list[str] = []
            for sheet_name in document.sheets:
                if sheet_name not in sheets:
                    unavailable_sheets.append(sheet_name)
                    continue
                sheet = workbook[sheet_name]
                has_data = any(
                    cell.value not in (None, "")
                    for row in sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row, 25),
                        min_col=1,
                        max_col=min(sheet.max_column, 25),
                    )
                    for cell in row
                )
                if has_data:
                    export_sheets.append(sheet_name)
                else:
                    unavailable_sheets.append(sheet_name)

            available = bool(export_sheets)
            if available and unavailable_sheets:
                reason = (
                    "Serão incluídos os separadores com dados: "
                    f"{', '.join(export_sheets)}. Sem dados: "
                    f"{', '.join(unavailable_sheets)}."
                )
            elif available:
                reason = f"Separadores incluídos: {', '.join(export_sheets)}."
            else:
                reason = (
                    "Nenhum dos separadores necessários contém dados: "
                    f"{', '.join(document.sheets)}."
                )
            result.append(
                PdfDocumentState(
                    document,
                    available,
                    reason,
                    tuple(export_sheets),
                )
            )
        return result
    finally:
        workbook.close()


def collision_free_path(folder: Path, filename: str) -> Path:
    base = Path(folder) / filename
    if not base.exists():
        return base
    for number in range(2, 10_000):
        candidate = base.with_name(f"{base.stem}_{number}{base.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Não foi possível criar um nome livre para {base.name}.")


def normalize_pdf_identifiers(identifiers: Iterable[str]) -> set[str]:
    """Mapeia presets antigos para o catálogo atual sem reativar documentos removidos."""
    selected = {str(identifier) for identifier in identifiers}
    if selected & FERRAGENS_LEGACY_IDS:
        selected.add("ferragens")
    selected.difference_update(RETIRED_DOCUMENT_IDS)
    return selected


def document_filename(document: PdfDocument, nome_enc_imos: str = "") -> str:
    """Resolve o nome final de um PDF, protegendo o fragmento vindo do Excel."""
    if "{nome_enc_imos}" not in document.filename:
        return document.filename
    nome = str(nome_enc_imos or "").strip()
    if not nome:
        raise ValueError(
            "Nome Enc IMOS IX em falta no Excel (DEFENICOES!E3)."
        )
    invalidos = '<>:"/\\|?*'
    nome_seguro = "".join(
        "_" if char in invalidos or ord(char) < 32 else char for char in nome
    )
    nome_seguro = nome_seguro.strip(" .")
    if not nome_seguro:
        raise ValueError("Nome Enc IMOS IX inválido para criar o nome do PDF.")
    return document.filename.format(nome_enc_imos=nome_seguro)


def read_nome_enc_imos_ix(workbook_path: Path) -> str:
    """Lê o Nome Enc IMOS IX do contrato estável DEFENICOES!E3."""
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        if "DEFENICOES" not in workbook.sheetnames:
            return ""
        return str(workbook["DEFENICOES"]["E3"].value or "").strip()
    finally:
        workbook.close()


def _unique_sheet_names(states: Iterable[PdfDocumentState]) -> tuple[str, ...]:
    result: list[str] = []
    for state in states:
        for sheet_name in state.export_sheets:
            if sheet_name not in result:
                result.append(sheet_name)
    return tuple(result)


def _export_sheets_to_pdf(
    excel, workbook, sheet_names: tuple[str, ...], output: Path
) -> None:
    if not sheet_names:
        raise ValueError("Não existem separadores com dados para exportar.")
    if len(sheet_names) == 1:
        workbook.Worksheets.Item(sheet_names[0]).ExportAsFixedFormat(0, str(output))
        return
    workbook.Worksheets.Item(sheet_names[0]).Select(True)
    for sheet_name in sheet_names[1:]:
        workbook.Worksheets.Item(sheet_name).Select(False)
    excel.ActiveSheet.ExportAsFixedFormat(0, str(output))


def export_pdf_documents(
    workbook_path: Path,
    destination: Path,
    identifiers: Iterable[str],
    *,
    export_separate: bool = True,
    create_package: bool = False,
    package_name: str = "Documentacao_Producao.pdf",
    progress_callback: Callable[[str, int, int], None] | None = None,
) -> PdfExportResult:
    selected_ids = normalize_pdf_identifiers(identifiers)
    states = {state.document.identifier: state for state in inspect_pdf_documents(workbook_path)}
    selected = [
        states[item.identifier]
        for item in DEFAULT_DOCUMENTS
        if item.identifier in selected_ids and item.identifier in states
    ]
    unavailable = [state for state in selected if not state.available]
    if unavailable:
        details = "\n".join(f"- {item.document.name}: {item.reason}" for item in unavailable)
        raise ValueError("Existem documentos indisponíveis:\n" + details)
    if not selected:
        raise ValueError("Selecione pelo menos um documento disponível.")

    nome_enc_imos = ""
    if any("{nome_enc_imos}" in state.document.filename for state in selected):
        nome_enc_imos = read_nome_enc_imos_ix(workbook_path)

    destination = Path(destination)
    destination.mkdir(parents=True, exist_ok=True)
    try:
        import win32com.client as win32_client
    except ImportError as exc:
        raise RuntimeError("A exportação necessita do Microsoft Excel e do pywin32.") from exc

    excel = None
    workbook = None
    outputs: list[Path] = []
    errors: list[str] = []
    package: Path | None = None
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass
        workbook = excel.Workbooks.Open(str(Path(workbook_path).resolve()), ReadOnly=True)
        total = len(selected)
        if export_separate:
            for index, state in enumerate(selected, start=1):
                document = state.document
                if progress_callback:
                    progress_callback(f"A exportar {document.name}…", index - 1, total)
                output = collision_free_path(
                    destination, document_filename(document, nome_enc_imos)
                )
                try:
                    _export_sheets_to_pdf(
                        excel, workbook, state.export_sheets, output
                    )
                    outputs.append(output)
                except Exception as exc:
                    errors.append(f"{document.name}: {exc}")
        if create_package:
            if progress_callback:
                progress_callback("A criar o pacote combinado…", total, total)
            package = collision_free_path(destination, package_name)
            _export_sheets_to_pdf(
                excel, workbook, _unique_sheet_names(selected), package
            )
        if progress_callback:
            progress_callback("Exportação concluída.", total, total)
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

    return PdfExportResult(tuple(outputs), package, tuple(errors))


class PdfPresetService:
    def __init__(self, session: Session) -> None:
        self.session = session

    def list(self, *, user_id: int, client: str) -> list[ListaMaterialPdfPreset]:
        key = _client_key(client)
        return list(
            self.session.execute(
                select(ListaMaterialPdfPreset)
                .where(
                    ListaMaterialPdfPreset.user_id == int(user_id),
                    ListaMaterialPdfPreset.cliente_chave == key,
                )
                .order_by(
                    ListaMaterialPdfPreset.predefinido.desc(),
                    ListaMaterialPdfPreset.ultimo_usado.desc(),
                    ListaMaterialPdfPreset.nome,
                )
            ).scalars()
        )

    def save(
        self,
        *,
        user_id: int,
        client: str,
        name: str,
        identifiers: Iterable[str],
        export_separate: bool,
        create_package: bool,
        make_default: bool = False,
    ) -> ListaMaterialPdfPreset:
        name = str(name or "").strip()
        if not name:
            raise ValueError("Indique um nome para o preset.")
        key = _client_key(client)
        self.session.execute(
            update(ListaMaterialPdfPreset)
            .where(
                ListaMaterialPdfPreset.user_id == int(user_id),
                ListaMaterialPdfPreset.cliente_chave == key,
            )
            .values(ultimo_usado=False)
        )
        if make_default:
            # Só pode existir um preset predefinido por utilizador/cliente.
            # A filtragem pelo user_id impede que esta escolha altere ou
            # exponha presets pertencentes a outro utilizador.
            self.session.execute(
                update(ListaMaterialPdfPreset)
                .where(
                    ListaMaterialPdfPreset.user_id == int(user_id),
                    ListaMaterialPdfPreset.cliente_chave == key,
                )
                .values(predefinido=False)
            )
        row = self.session.execute(
            select(ListaMaterialPdfPreset).where(
                ListaMaterialPdfPreset.user_id == int(user_id),
                ListaMaterialPdfPreset.cliente_chave == key,
                ListaMaterialPdfPreset.nome == name,
            )
        ).scalar_one_or_none()
        if row is None:
            row = ListaMaterialPdfPreset(user_id=int(user_id), cliente_chave=key, nome=name)
            self.session.add(row)
        row.documentos_json = json.dumps(list(identifiers), ensure_ascii=False)
        row.exportar_separados = bool(export_separate)
        row.criar_pacote = bool(create_package)
        row.ultimo_usado = True
        row.predefinido = bool(make_default)
        self.session.commit()
        return row


def _client_key(value: str) -> str:
    from app.services.lista_material_assistente_service import client_key

    return client_key(value)
