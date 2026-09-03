"""Export the raw-material catalog to Excel, for reading and printing.

The V3 owns the catalog now; this is the way out — a file to consult, to print
or to send to someone, never a file to import back. Nothing here writes to the
database, and the export deliberately carries the same columns the user has on
screen: what they exported is what they were looking at.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

_FMT_EUR = "#,##0.00 €"
_FMT_PCT = "0.0\\%"
_FMT_NUM = "0.###"
_FMT_DATA = "dd-mm-yyyy"

COR_CABECALHO = "8B6F4E"
COR_ZEBRA = "F7F2EA"
COR_INATIVO = "9A928A"

#: Colunas cujo conteúdo é dinheiro, percentagem ou data, para o Excel as
#: formatar como tal em vez de as deixar como texto.
FORMATOS = {
    "Preço tabela": _FMT_EUR,
    "Preço Líquido": _FMT_EUR,
    "Desc %": _FMT_PCT,
    "Mrg %": _FMT_PCT,
    "Desp %": _FMT_PCT,
    "Último preço": _FMT_DATA,
    "Comp MP": _FMT_NUM,
    "Larg MP": _FMT_NUM,
    "Esp MP": _FMT_NUM,
}

LARGURAS = {
    "Ref LE": 11,
    "Descrição": 48,
    "Tipo Excel": 18,
    "Família Excel": 14,
    "Fornecedor": 18,
    "Ref. fornecedor": 20,
    "Fabricante": 16,
    "Link": 46,
    "Imagem": 34,
    "Observações": 32,
}
LARGURA_PADRAO = 13


def nome_do_ficheiro(hoje: date | None = None) -> str:
    """Nome sugerido para o ficheiro exportado."""
    return f"Materias_Primas_{(hoje or date.today()):%Y-%m-%d}.xlsx"


def exportar_materias_primas(
    colunas,
    linhas,
    caminho: str | Path,
    titulo: str = "Matérias-Primas — Martelo V3",
    hoje: date | None = None,
) -> Path:
    """Escrever o catálogo num ficheiro Excel.

    ``linhas`` são tuplos ``(valores, ativo)``: os valores já vêm prontos para
    escrever (na mesma ordem das colunas) e o ``ativo`` serve só para as
    descontinuadas saírem a cinzento, como aparecem no ecrã.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    caminho = Path(caminho)
    livro = openpyxl.Workbook()
    folha = livro.active
    folha.title = "Materias-Primas"

    folha.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)
    folha.cell(
        row=1,
        column=max(len(colunas), 2),
        value=f"Exportado em {(hoje or date.today()):%d-%m-%Y}",
    ).alignment = Alignment(horizontal="right")

    for indice, coluna in enumerate(colunas, start=1):
        celula = folha.cell(row=2, column=indice, value=coluna)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        folha.column_dimensions[get_column_letter(indice)].width = LARGURAS.get(
            coluna, LARGURA_PADRAO
        )

    for numero, (valores, ativo) in enumerate(linhas, start=3):
        zebra = numero % 2 == 0
        for indice, (coluna, valor) in enumerate(zip(colunas, valores), start=1):
            celula = folha.cell(row=numero, column=indice)
            celula.value = float(valor) if isinstance(valor, Decimal) else valor
            formato = FORMATOS.get(coluna)
            if formato and not isinstance(valor, str):
                celula.number_format = formato
            if zebra:
                celula.fill = PatternFill("solid", fgColor=COR_ZEBRA)
            if not ativo:
                # Descontinuada: riscada e a cinzento, como no ecrã.
                celula.font = Font(strike=True, color=COR_INATIVO)

    folha.freeze_panes = "A3"
    if colunas and linhas:
        ultima = get_column_letter(len(colunas))
        folha.auto_filter.ref = f"A2:{ultima}{len(linhas) + 2}"

    caminho.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho)

    return caminho
