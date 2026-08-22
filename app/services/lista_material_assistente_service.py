"""Motor assistido IMOS → CUT-RITE.

O motor nunca consulta ``def_materias_primas``. Sugestões e operações CNC são
separadas, explicáveis e permanecem pendentes até decisão humana.
"""

from __future__ import annotations

import json
import importlib
import math
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from app.models.lista_material_assistente import (
    ListaMaterialCncOperacao,
    ListaMaterialBarraReceita,
    ListaMaterialAliasPlaca,
    ListaMaterialExecucao,
    ListaMaterialModulo,
    ListaMaterialObraConfig,
    ListaMaterialPerfil,
    ListaMaterialRelacaoOrla,
    ListaMaterialSugestao,
)
from app.services.warehouse_board_catalog import (
    BoardRecord,
    BoardCatalogProvider,
    UnavailableBoardCatalogProvider,
)


SHEET_CUTRITE = "LISTAGEM_CUT_RITE"
EDGE_FIELDS = ("Orla ESQ", "Orla DIR", "Orla CIMA", "Orla BAIXO")
MODULES = (
    "vista_vertical",
    "remate_teto",
    "rodape_frente",
    "agrupamentos",
    "cnc_fresar",
    "notas",
    "puxadores",
    "lacagem_formal",
    "validacao_placa_orla",
    "sugestoes_material",
    "exportacao_pdf",
)
DEFAULT_MODULES = {name: True for name in MODULES} | {"lacagem_formal": False}
LACQUER_BOARD_KEY = "MDF_MR_MLM_BRANCO_B3002_MA_19MM"
DOOR_LACQUER_DESCRIPTIONS = (
    "PORTA_DIREITA",
    "PORTA_ESQUERDA",
    "PORTA_SUBIR_DESCER",
    "FRENTE_DE_GAVETA",
)
TETO_FUNDO_DESCRIPTIONS = {"TETO", "FUNDO"}
MALEIRO_PRATELEIRA_DESCRIPTIONS = ("MALEIRO", "PRATELEIRA_AMOVIVEL")
LATERAL_DESCRIPTIONS = {"LATERAL_ESQUERDA", "LATERAL_DIREITA"}


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_")


def client_key(value: object) -> str:
    return normalize_text(value)[:160]


def _decimal(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def extract_edge_thickness(edge_name: str) -> Decimal | None:
    match = re.search(r"(?:PVC|ABS)[ _-]?(\d+)[.,](\d+)", str(edge_name or ""), re.I)
    if not match:
        return None
    return Decimal(f"{match.group(1)}.{match.group(2)}")


def extract_handle(description: str) -> str:
    """Extrai uma sugestão conservadora; nunca obriga a aplicá-la."""
    text = " ".join(str(description or "").split())
    patterns = (
        r"(?i)\bpuxador\s*[:\-]?\s*([^\n;,.]{2,60})",
        r"(?i)\btic[\s-]*tac\b",
        r"(?i)\bpuxador\s+j\s*([A-Z]?\d{3,5})?",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0).strip(" :-")[:255]
    return ""


def compact_handle(value: str) -> str:
    """Converte o puxador para uma nota curta, adequada à etiqueta."""
    text = " ".join(str(value or "").strip().split())
    text = re.sub(r"(?i)^puxador\s*", "", text).strip(" :-")
    text = re.sub(r"(?i)^pux\s*", "", text).strip(" :-")
    return f"Pux {text}".strip() if text else ""


def _remove_note_fragments(value: str, fragments: Iterable[str]) -> tuple[str, bool]:
    """Remove apenas fragmentos conhecidos, preservando as restantes notas locais."""
    parts = [part.strip() for part in str(value or "").split(";")]
    cleaned_parts: list[str] = []
    changed = False
    usable = sorted(
        {str(fragment).strip() for fragment in fragments if str(fragment).strip()},
        key=len,
        reverse=True,
    )
    for part in parts:
        cleaned = part
        for fragment in usable:
            updated = re.sub(re.escape(fragment), "", cleaned, flags=re.IGNORECASE)
            if updated != cleaned:
                changed = True
                cleaned = updated
        cleaned = re.sub(r"\s*\+\s*$", "", cleaned).strip(" +")
        if cleaned:
            cleaned_parts.append(cleaned)
    return "; ".join(cleaned_parts), changed


def _description_is(value: str, keys: Iterable[str]) -> bool:
    description = normalize_text(value)
    return any(key in description for key in keys)


@dataclass(frozen=True)
class AssistantConfig:
    user_id: int
    client: str
    modules: dict[str, bool] = field(default_factory=lambda: dict(DEFAULT_MODULES))
    handle: str = ""
    handle_exceptions: dict[str, str] = field(default_factory=dict)
    cnc_note: str = "CNC_FRESAR"
    formal_lacquering: bool = False
    board_catalog_available: bool = False
    board_catalog_message: str = ""

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class MaterialRow:
    row_number: int
    source_id: str
    description: str
    material: str
    length: Decimal | None
    width: Decimal | None
    quantity: Decimal | None
    article: str
    notes: str
    edges: dict[str, str]
    values: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class AssistantSuggestion:
    source_id: str
    row_number: int | None
    field: str
    original: str
    suggested: str
    reason: str
    confidence: float
    kind: str
    blocking: bool = False
    allow_blank: bool = False
    delete_row: bool = False
    group_id: str = ""


@dataclass(frozen=True)
class WorkbookAudit:
    workbook_path: Path
    rows: tuple[MaterialRow, ...]
    suggestions: tuple[AssistantSuggestion, ...]
    blocking: tuple[AssistantSuggestion, ...]
    board_catalog_message: str
    columns: tuple[str, ...] = ()


@dataclass(frozen=True)
class AssistantDecision:
    suggestion: AssistantSuggestion
    action: str  # aceitar, rejeitar ou editar
    value: str = ""


@dataclass(frozen=True)
class BoardMatch:
    board: BoardRecord | None
    confidence: float
    reason: str
    status: str


class ListaMaterialAssistantService:
    def __init__(
        self,
        session: Session,
        *,
        board_catalog: BoardCatalogProvider | None = None,
    ) -> None:
        self.session = session
        self.board_catalog = board_catalog or UnavailableBoardCatalogProvider()
        self._known_edges_cache: dict[tuple[str, int, str], list[str]] = {}
        self._edge_relation_cache: dict[
            tuple[str, str, int, str], ListaMaterialRelacaoOrla
        ] = {}

    def resolve_config(
        self,
        *,
        user_id: int,
        client: str,
        production_description: str = "",
    ) -> AssistantConfig:
        key = client_key(client)
        modules = dict(DEFAULT_MODULES)
        # Precedência: global -> cliente -> utilizador -> utilizador+cliente.
        rows = list(
            self.session.execute(
                select(ListaMaterialModulo).where(
                    or_(ListaMaterialModulo.user_id.in_((0, int(user_id))),),
                    or_(ListaMaterialModulo.cliente_chave.in_(("", key)),),
                )
            ).scalars()
        )
        rows.sort(key=lambda row: (bool(row.user_id), bool(row.cliente_chave)))
        for row in rows:
            if row.modulo in modules:
                modules[row.modulo] = bool(row.ativo)

        profile = self.session.execute(
            select(ListaMaterialPerfil).where(
                ListaMaterialPerfil.user_id == int(user_id),
                ListaMaterialPerfil.cliente_chave == key,
            )
        ).scalar_one_or_none()
        handle = (profile.puxador_default if profile else "") or extract_handle(production_description)
        handle_exceptions: dict[str, str] = {}
        cnc_note = "CNC_FRESAR"
        saved_config: dict[str, object] = {}
        if profile and profile.configuracao_json:
            try:
                saved_config = json.loads(profile.configuracao_json)
                handle_exceptions = {
                    str(key): str(value)
                    for key, value in dict(saved_config.get("handle_exceptions") or {}).items()
                }
            except (TypeError, ValueError, json.JSONDecodeError):
                handle_exceptions = {}
                saved_config = {}
        if "cnc_note" in saved_config:
            cnc_note = str(saved_config.get("cnc_note") or "")
        for exception_key in list(handle_exceptions):
            if normalize_text(exception_key) == "CNC_FRESAR":
                if "cnc_note" not in saved_config:
                    cnc_note = handle_exceptions[exception_key]
                handle_exceptions.pop(exception_key)
        formal = bool(profile.lacagem_formal) if profile else False
        # Regra confirmada para Paulo + JF_VIVA.
        if normalize_text(client) == "JF_VIVA":
            formal = False
            modules["lacagem_formal"] = False
        status = self.board_catalog.status()
        return AssistantConfig(
            user_id=int(user_id),
            client=client,
            modules=modules,
            handle=handle,
            handle_exceptions=handle_exceptions,
            cnc_note=cnc_note,
            formal_lacquering=formal,
            board_catalog_available=status.available,
            board_catalog_message=status.message,
        )

    def resolve_work_config(
        self,
        *,
        production_id: int,
        user_id: int,
        client: str,
        production_description: str = "",
    ) -> AssistantConfig:
        """Recupera a configuração escolhida ao criar o livro desta obra.

        Se não existir snapshot (livros antigos), aplica as preferências atuais
        do utilizador/cliente. O estado do catálogo é sempre consultado de novo.
        """
        snapshot = self.session.execute(
            select(ListaMaterialObraConfig)
            .where(
                ListaMaterialObraConfig.producao_id == int(production_id),
                ListaMaterialObraConfig.user_id == int(user_id),
            )
            .order_by(ListaMaterialObraConfig.id.desc())
        ).scalars().first()
        if snapshot is None:
            return self.resolve_config(
                user_id=user_id,
                client=client,
                production_description=production_description,
            )

        try:
            saved = json.loads(snapshot.configuracao_json or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            saved = {}
        try:
            exceptions = json.loads(snapshot.puxadores_excecoes_json or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            exceptions = {}
        cnc_note = (
            str(saved.get("cnc_note") or "")
            if "cnc_note" in saved
            else "CNC_FRESAR"
        )
        for exception_key in list(exceptions):
            if normalize_text(exception_key) == "CNC_FRESAR":
                if "cnc_note" not in saved:
                    cnc_note = str(exceptions[exception_key])
                exceptions.pop(exception_key)
        modules = dict(DEFAULT_MODULES)
        modules.update(
            {
                str(name): bool(enabled)
                for name, enabled in dict(saved.get("modules") or {}).items()
                if str(name) in modules
            }
        )
        status = self.board_catalog.status()
        return AssistantConfig(
            user_id=int(user_id),
            client=client,
            modules=modules,
            handle=str(snapshot.puxador_obra or saved.get("handle") or ""),
            handle_exceptions={
                str(key): str(value) for key, value in dict(exceptions or {}).items()
            },
            cnc_note=cnc_note,
            formal_lacquering=bool(saved.get("formal_lacquering", False)),
            board_catalog_available=status.available,
            board_catalog_message=status.message,
        )

    def identify_board(
        self,
        material: str,
        *,
        thickness: Decimal | None,
        config: AssistantConfig,
    ) -> BoardMatch:
        material_key = normalize_text(material)
        status = self.board_catalog.status()
        if not status.available:
            return BoardMatch(None, 0.0, status.message, "desconhecido")
        boards = self.board_catalog.list_boards()
        by_id = {board.external_id: board for board in boards}
        aliases = list(
            self.session.execute(
                select(ListaMaterialAliasPlaca).where(
                    ListaMaterialAliasPlaca.texto_normalizado == material_key,
                    ListaMaterialAliasPlaca.estado == "aprovado",
                    ListaMaterialAliasPlaca.user_id.in_((0, config.user_id)),
                    ListaMaterialAliasPlaca.cliente_chave.in_(("", client_key(config.client))),
                )
            ).scalars()
        )
        aliases.sort(
            key=lambda item: (bool(item.user_id), bool(item.cliente_chave), item.suporte),
            reverse=True,
        )
        for alias in aliases:
            if alias.placa_externa_id in by_id:
                return BoardMatch(
                    by_id[alias.placa_externa_id],
                    float(alias.confianca),
                    "Alias explicitamente aprovado para este âmbito.",
                    "confirmado",
                )

        scored: list[tuple[float, BoardRecord, str]] = []
        material_tokens = set(material_key.split("_"))
        for board in boards:
            code_key = normalize_text(board.code)
            description_key = normalize_text(board.description)
            if material_key == code_key:
                score, reason = 1.0, "Código/referência HOMAG exato."
            elif code_key and (code_key in material_key or material_key in code_key):
                score, reason = 0.95, "Referência HOMAG contida no texto do material."
            else:
                board_tokens = set(description_key.split("_")) | set(code_key.split("_"))
                union = material_tokens | board_tokens
                score = len(material_tokens & board_tokens) / len(union) if union else 0.0
                reason = "Semelhança entre descrição, cor, marca, acabamento e textura."
            if thickness is not None and board.thickness is not None:
                if abs(board.thickness - thickness) <= Decimal("0.2"):
                    score = min(1.0, score + 0.08)
                    reason += " Espessura compatível."
                else:
                    score = max(0.0, score - 0.25)
                    reason += " Espessura diferente."
            scored.append((score, board, reason))
        scored.sort(key=lambda item: item[0], reverse=True)
        if not scored or scored[0][0] < 0.45:
            return BoardMatch(None, scored[0][0] if scored else 0.0, "Placa ainda não identificada no armazém.", "desconhecido")
        best = scored[0]
        status_name = "provavel" if best[0] < 0.90 else "confirmado"
        return BoardMatch(best[1], best[0], best[2], status_name)

    def save_profile_defaults(self, config: AssistantConfig) -> None:
        key = client_key(config.client)
        profile = self.session.execute(
            select(ListaMaterialPerfil).where(
                ListaMaterialPerfil.user_id == config.user_id,
                ListaMaterialPerfil.cliente_chave == key,
            )
        ).scalar_one_or_none()
        if profile is None:
            profile = ListaMaterialPerfil(user_id=config.user_id, cliente_chave=key)
            self.session.add(profile)
        profile.puxador_default = config.handle or None
        profile.lacagem_formal = bool(config.formal_lacquering)
        profile.configuracao_json = json.dumps(config.to_dict(), ensure_ascii=False)
        for name, enabled in config.modules.items():
            module = self.session.execute(
                select(ListaMaterialModulo).where(
                    ListaMaterialModulo.user_id == config.user_id,
                    ListaMaterialModulo.cliente_chave == key,
                    ListaMaterialModulo.modulo == name,
                )
            ).scalar_one_or_none()
            if module is None:
                module = ListaMaterialModulo(
                    user_id=config.user_id,
                    cliente_chave=key,
                    modulo=name,
                )
                self.session.add(module)
            module.ativo = bool(enabled)
        self.session.commit()

    def create_work_snapshot(
        self,
        *,
        production_id: int,
        workbook_path: Path,
        config: AssistantConfig,
        handle_exceptions: Mapping[str, str] | None = None,
    ) -> ListaMaterialObraConfig:
        row = ListaMaterialObraConfig(
            producao_id=int(production_id),
            user_id=config.user_id,
            cliente_chave=client_key(config.client),
            workbook_path=str(workbook_path),
            puxador_obra=config.handle or None,
            puxadores_excecoes_json=json.dumps(
                dict(config.handle_exceptions if handle_exceptions is None else handle_exceptions),
                ensure_ascii=False,
            ),
            modulos_json=json.dumps(config.modules, ensure_ascii=False),
            configuracao_json=json.dumps(config.to_dict(), ensure_ascii=False),
        )
        self.session.add(row)
        self.session.commit()
        return row

    def audit_workbook(
        self,
        workbook_path: Path,
        *,
        config: AssistantConfig,
    ) -> WorkbookAudit:
        columns, material_rows = read_material_table(workbook_path)
        rows = tuple(material_rows)
        suggestions = tuple(self.analyze_rows(rows, config=config))
        return WorkbookAudit(
            workbook_path=Path(workbook_path),
            rows=rows,
            suggestions=suggestions,
            blocking=tuple(item for item in suggestions if item.blocking),
            board_catalog_message=config.board_catalog_message,
            columns=columns,
        )

    def analyze_rows(
        self,
        rows: Iterable[MaterialRow],
        *,
        config: AssistantConfig,
    ) -> list[AssistantSuggestion]:
        rows = list(rows)
        result: list[AssistantSuggestion] = []
        if config.modules.get("cnc_fresar", True):
            for row in rows:
                result.extend(self._cnc_suggestions(row, config))
            result.extend(self._teto_fundo_edge_suggestions(rows))
        if config.modules.get("notas", True):
            result.extend(self._note_suggestions(rows, config))
        if config.modules.get("vista_vertical", True):
            result.extend(_vista_vertical_suggestions(rows))
        if config.modules.get("remate_teto", True):
            result.extend(
                _threshold_bar_suggestions(
                    rows,
                    description_key="REMATE_TETO",
                    target_width=Decimal("75"),
                    split_threshold=Decimal("2780"),
                    split_length=Decimal("2780"),
                    medium_threshold=Decimal("2070"),
                    short_length=Decimal("2050"),
                )
            )
            result.extend(self._remate_lacquer_conflicts(rows, config))
        if config.modules.get("rodape_frente", True):
            result.extend(
                _threshold_bar_suggestions(
                    rows,
                    description_key="RODAPE_FRENTE",
                    target_width=Decimal("75"),
                    split_threshold=Decimal("2850"),
                    split_length=Decimal("2830"),
                    medium_threshold=Decimal("2100"),
                    short_length=Decimal("2070"),
                )
            )
        return result

    def _cnc_suggestions(
        self, row: MaterialRow, config: AssistantConfig
    ) -> list[AssistantSuggestion]:
        suggestions: list[AssistantSuggestion] = []
        for side, edge in row.edges.items():
            if "CNC_FRESAR" not in normalize_text(edge):
                continue
            description_key = normalize_text(row.description)
            if description_key in LATERAL_DESCRIPTIONS:
                suggestions.append(
                    AssistantSuggestion(
                        source_id=row.source_id,
                        row_number=row.row_number,
                        field=side,
                        original=edge,
                        suggested="",
                        reason=(
                            f"{row.description}: CNC_FRESAR é limpo apenas deste lado; "
                            "as restantes orlas permanecem inalteradas."
                        ),
                        confidence=0.99,
                        kind="cnc_fresar_lateral_vazio",
                        blocking=False,
                        allow_blank=True,
                    )
                )
                continue
            if description_key in TETO_FUNDO_DESCRIPTIONS and side != "Orla ESQ":
                suggestions.append(
                    AssistantSuggestion(
                        source_id=row.source_id,
                        row_number=row.row_number,
                        field=side,
                        original=edge,
                        suggested="",
                        reason=(
                            f"{row.description}: apenas Orla ESQ é aplicável; "
                            "CNC_FRESAR mantém-se como operação nas notas."
                        ),
                        confidence=0.97,
                        kind="cnc_fresar_teto_fundo",
                        allow_blank=True,
                    )
                )
                continue
            if _description_is(
                row.description, MALEIRO_PRATELEIRA_DESCRIPTIONS
            ):
                candidate_sides = ("Orla DIR", "Orla CIMA", "Orla BAIXO")
            else:
                candidate_sides = EDGE_FIELDS
            alternatives = sorted(
                {
                    row.edges.get(other, "")
                    for other in candidate_sides
                    if other != side
                    and row.edges.get(other, "")
                    and "CNC_FRESAR"
                    not in normalize_text(row.edges.get(other, ""))
                }
            )
            if len(alternatives) == 1:
                proposed = alternatives[0]
                reason = "CNC_FRESAR foi separado como operação; existe uma única orla nos restantes lados."
                confidence = 0.90
                blocking = False
            else:
                known = self._known_edges(row.material, config)
                proposed = known[0] if len(known) == 1 else ""
                if proposed:
                    reason = "CNC_FRESAR foi separado como operação; única relação placa–orla aprovada no histórico."
                    confidence = 0.82
                    blocking = False
                else:
                    reason = "CNC_FRESAR é uma operação, não uma orla; é necessário escolher a orla deste lado."
                    confidence = 0.35
                    blocking = True
            suggestions.append(
                AssistantSuggestion(
                    source_id=row.source_id,
                    row_number=row.row_number,
                    field=side,
                    original=edge,
                    suggested=proposed,
                    reason=reason,
                    confidence=confidence,
                    kind="cnc_fresar",
                    blocking=blocking,
                )
            )
        return suggestions

    @staticmethod
    def _teto_fundo_edge_suggestions(
        rows: Iterable[MaterialRow],
    ) -> list[AssistantSuggestion]:
        result: list[AssistantSuggestion] = []
        for row in rows:
            if normalize_text(row.description) not in TETO_FUNDO_DESCRIPTIONS:
                continue
            for side in ("Orla DIR", "Orla CIMA", "Orla BAIXO"):
                edge = row.edges.get(side, "")
                if not edge or "CNC_FRESAR" in normalize_text(edge):
                    continue
                result.append(
                    AssistantSuggestion(
                        source_id=row.source_id,
                        row_number=row.row_number,
                        field=side,
                        original=edge,
                        suggested="",
                        reason=f"{row.description}: apenas a coluna Orla ESQ deve ficar preenchida.",
                        confidence=0.97,
                        kind="teto_fundo_orla_vazia",
                        allow_blank=True,
                    )
                )
        return result

    def _note_suggestions(
        self, rows: Iterable[MaterialRow], config: AssistantConfig
    ) -> list[AssistantSuggestion]:
        result: list[AssistantSuggestion] = []
        for row in rows:
            additions: list[str] = []
            reasons: list[str] = []
            description_key = normalize_text(row.description)
            material_key = normalize_text(row.material)
            has_cnc = any(
                "CNC_FRESAR" in normalize_text(edge)
                for edge in row.edges.values()
            )
            if description_key in LATERAL_DESCRIPTIONS and has_cnc:
                if row.notes.strip():
                    result.append(
                        AssistantSuggestion(
                            source_id=row.source_id,
                            row_number=row.row_number,
                            field="Notas",
                            original=row.notes,
                            suggested="",
                            reason=(
                                f"{row.description} com CNC_FRESAR: o campo Notas "
                                "deve ficar vazio."
                            ),
                            confidence=0.99,
                            kind="cnc_fresar_lateral_notas_vazias",
                            blocking=False,
                            allow_blank=True,
                        )
                    )
                continue
            raw_handle = ""
            if config.modules.get("puxadores", True):
                article_key = normalize_text(row.article)
                for key, value in config.handle_exceptions.items():
                    if normalize_text(key) == article_key:
                        raw_handle = value
                        break
                if not raw_handle:
                    raw_handle = config.handle
            handle = compact_handle(raw_handle)
            base_notes, removed_handle = _remove_note_fragments(
                row.notes, (raw_handle, handle)
            )
            if removed_handle:
                reasons.append(
                    "puxador anterior removido/reformatado segundo o tipo de peça"
                )
            cnc_note = str(config.cnc_note or "").strip()
            removed_legacy_cnc = False
            if has_cnc and normalize_text(cnc_note) != "CNC_FRESAR":
                base_notes, removed_legacy_cnc = _remove_note_fragments(
                    base_notes, ("CNC_FRESAR",)
                )
                if removed_legacy_cnc:
                    reasons.append("indicação CNC anterior substituída pela configuração da obra")
            if (
                has_cnc
                and cnc_note
                and normalize_text(cnc_note) not in normalize_text(base_notes)
            ):
                additions.append(cnc_note)
                reasons.append(
                    "operação CNC detetada num campo de orla; texto definido na configuração da obra"
                )
            has_lacquer = any("LACAR" in normalize_text(edge) for edge in row.edges.values())
            lacquer_piece = _description_is(
                row.description, DOOR_LACQUER_DESCRIPTIONS
            )
            if has_lacquer and lacquer_piece:
                base_notes, removed_lacquer = _remove_note_fragments(
                    base_notes, ("Lacar 1 Face",)
                )
                if removed_lacquer:
                    reasons.append("nota de lacagem compactada")
                finish_note = "Lacar 1 Face"
                if handle:
                    finish_note += f" + {handle}"
                additions.append(finish_note)
                material_reason = (
                    " e material de referência B3002/MA"
                    if material_key == LACQUER_BOARD_KEY
                    else ""
                )
                reasons.append(
                    "peça de porta/frente com orla LACAR" + material_reason
                )
            elif "PORTA" in description_key and handle:
                additions.append(handle)
                reasons.append("puxador aplicado apenas a uma peça de porta")
            if (
                "REMATE_TETO" in description_key
                and material_key == LACQUER_BOARD_KEY
                and "NAO_LACAR" not in normalize_text(base_notes)
            ):
                additions.append("Não Lacar")
                reasons.append("Remate Teto B3002/MA nasce como não lacar")
            if not additions and not (removed_handle or removed_legacy_cnc):
                continue
            suggested = base_notes
            for addition in additions:
                if normalize_text(addition) not in normalize_text(suggested):
                    suggested = f"{suggested}; {addition}" if suggested else addition
            if suggested == row.notes.strip():
                continue
            result.append(
                AssistantSuggestion(
                    source_id=row.source_id,
                    row_number=row.row_number,
                    field="Notas",
                    original=row.notes,
                    suggested=suggested,
                    reason="; ".join(reasons) + ".",
                    confidence=0.92,
                    kind="notas_assistente",
                    allow_blank=not suggested,
                )
            )
        return result

    @staticmethod
    def _remate_lacquer_conflicts(
        rows: Iterable[MaterialRow], config: AssistantConfig
    ) -> list[AssistantSuggestion]:
        if client_key(config.client) != "JF_VIVA":
            return []
        result: list[AssistantSuggestion] = []
        for row in rows:
            if "REMATE_TETO" not in normalize_text(row.description):
                continue
            for side, edge in row.edges.items():
                if "LACAR" not in normalize_text(edge):
                    continue
                result.append(
                    AssistantSuggestion(
                        source_id=row.source_id,
                        row_number=row.row_number,
                        field=side,
                        original=edge,
                        suggested="",
                        reason=(
                            "Conflito: para Paulo + JF_VIVA o Remate Teto nasce como não lacar. "
                            "Confirme localmente se esta obra é uma exceção."
                        ),
                        confidence=0.80,
                        kind="remate_teto_lacagem",
                        allow_blank=True,
                    )
                )
        return result

    def _known_edges(self, material: str, config: AssistantConfig) -> list[str]:
        key = normalize_text(material)
        cache_key = (key, config.user_id, client_key(config.client))
        if cache_key in self._known_edges_cache:
            return list(self._known_edges_cache[cache_key])
        rows = self.session.execute(
            select(ListaMaterialRelacaoOrla).where(
                ListaMaterialRelacaoOrla.material_normalizado == key,
                ListaMaterialRelacaoOrla.estado == "aprovado",
                or_(ListaMaterialRelacaoOrla.user_id.in_((0, config.user_id))),
                or_(ListaMaterialRelacaoOrla.cliente_chave.in_(("", client_key(config.client)))),
            ).order_by(ListaMaterialRelacaoOrla.confianca.desc(), ListaMaterialRelacaoOrla.suporte.desc())
        ).scalars()
        result = [row.orla_normalizada for row in rows]
        self._known_edges_cache[cache_key] = result
        return list(result)

    def _recipe_length_suggestions(
        self,
        rows: Iterable[MaterialRow],
        description_key: str,
        config: AssistantConfig,
    ) -> list[AssistantSuggestion]:
        candidates = list(
            self.session.execute(
                select(ListaMaterialBarraReceita).where(
                    ListaMaterialBarraReceita.descricao_normalizada == description_key,
                    ListaMaterialBarraReceita.estado == "aprovado",
                    ListaMaterialBarraReceita.user_id.in_((0, config.user_id)),
                    ListaMaterialBarraReceita.cliente_chave.in_(("", client_key(config.client))),
                )
            ).scalars()
        )
        result: list[AssistantSuggestion] = []
        for row in rows:
            if description_key not in normalize_text(row.description):
                continue
            material_key = normalize_text(row.material)
            applicable = [
                recipe
                for recipe in candidates
                if not recipe.material_normalizado or recipe.material_normalizado == material_key
            ]
            applicable.sort(
                key=lambda recipe: (
                    bool(recipe.user_id),
                    bool(recipe.cliente_chave),
                    bool(recipe.material_normalizado),
                    recipe.suporte,
                ),
                reverse=True,
            )
            if not applicable or applicable[0].comprimento is None:
                continue
            recipe = applicable[0]
            if row.length == recipe.comprimento:
                continue
            result.append(
                AssistantSuggestion(
                    source_id=row.source_id,
                    row_number=row.row_number,
                    field="Comp",
                    original=str(row.length or ""),
                    suggested=str(recipe.comprimento),
                    reason=(
                        f"Receita histórica/manual aprovada ({recipe.suporte} utilizações). "
                        "Não é apresentada como dimensão HOMAG."
                    ),
                    confidence=min(0.95, 0.60 + 0.03 * recipe.suporte),
                    kind="barra_comprimento_historico",
                )
            )
        return result

    def record_audit(
        self,
        *,
        production_id: int,
        user_id: int,
        audit: WorkbookAudit,
        kind: str = "pre_cutrite",
    ) -> ListaMaterialExecucao:
        execution = ListaMaterialExecucao(
            producao_id=int(production_id),
            user_id=int(user_id),
            tipo=kind,
            estado="bloqueada" if audit.blocking else "concluida",
            resumo_json=json.dumps(
                {
                    "workbook": str(audit.workbook_path),
                    "rows": len(audit.rows),
                    "suggestions": len(audit.suggestions),
                    "blocking": len(audit.blocking),
                    "board_catalog": audit.board_catalog_message,
                },
                ensure_ascii=False,
            ),
            concluida_em=datetime.now(),
        )
        self.session.add(execution)
        self.session.flush()
        for item in audit.suggestions:
            self.session.add(
                ListaMaterialSugestao(
                    execucao_id=execution.id,
                    source_id=item.source_id,
                    folha=SHEET_CUTRITE,
                    linha=item.row_number,
                    campo=item.field,
                    original=item.original,
                    sugerido=item.suggested,
                    motivo=item.reason,
                    confianca=Decimal(str(item.confidence)),
                )
            )
            if item.kind == "cnc_fresar":
                self.session.add(
                    ListaMaterialCncOperacao(
                        execucao_id=execution.id,
                        source_id=item.source_id,
                        lado=item.field,
                        orla_original=item.original,
                        orla_resolvida=item.suggested or None,
                        estado="proposta" if item.suggested else "pendente",
                    )
                )
        self.session.commit()
        return execution

    def record_decisions(
        self,
        *,
        execution_id: int,
        decisions: Iterable[AssistantDecision],
        rows: Iterable[MaterialRow],
        config: AssistantConfig,
    ) -> None:
        material_by_source = {row.source_id: row.material for row in rows}
        for decision in decisions:
            suggestion = decision.suggestion
            stored = self.session.execute(
                select(ListaMaterialSugestao).where(
                    ListaMaterialSugestao.execucao_id == int(execution_id),
                    ListaMaterialSugestao.source_id == suggestion.source_id,
                    ListaMaterialSugestao.campo == suggestion.field,
                )
            ).scalars().first()
            if stored is not None:
                stored.estado = decision.action
                stored.sugerido = decision.value or suggestion.suggested
                stored.decidido_por_id = config.user_id
                stored.decidido_em = datetime.now()
            if (
                decision.action in {"aceitar", "editar"}
                and suggestion.field in EDGE_FIELDS
                and (decision.value or suggestion.suggested)
            ):
                self._learn_edge_relation(
                    material_by_source.get(suggestion.source_id, ""),
                    decision.value or suggestion.suggested,
                    config,
                )
            elif (
                decision.action == "rejeitar"
                and suggestion.field in EDGE_FIELDS
                and suggestion.suggested
            ):
                self._learn_edge_rejection(
                    material_by_source.get(suggestion.source_id, ""),
                    suggestion.suggested,
                    config,
                )
        execution = self.session.get(ListaMaterialExecucao, int(execution_id))
        if execution is not None:
            execution.estado = "concluida"
            execution.concluida_em = datetime.now()
        self.session.commit()

    def _learn_edge_relation(self, material: str, edge: str, config: AssistantConfig) -> None:
        material_key = normalize_text(material)
        edge_key = normalize_text(edge)
        if not material_key or not edge_key or edge_key == "CNC_FRESAR":
            return
        relation_key = (
            material_key,
            edge_key,
            config.user_id,
            client_key(config.client),
        )
        row = self._edge_relation_cache.get(relation_key)
        if row is None:
            row = self.session.execute(
                select(ListaMaterialRelacaoOrla).where(
                    ListaMaterialRelacaoOrla.material_normalizado == material_key,
                    ListaMaterialRelacaoOrla.orla_normalizada == edge_key,
                    ListaMaterialRelacaoOrla.user_id == config.user_id,
                    ListaMaterialRelacaoOrla.cliente_chave == client_key(config.client),
                )
            ).scalar_one_or_none()
        if row is None:
            row = ListaMaterialRelacaoOrla(
                material_normalizado=material_key,
                orla_normalizada=edge_key,
                espessura_orla=extract_edge_thickness(edge),
                user_id=config.user_id,
                cliente_chave=client_key(config.client),
                origem="manual",
                estado="aprovado",
                confianca=Decimal("1"),
                suporte=1,
                confirmacoes=1,
                confirmado_por_id=config.user_id,
                primeira_utilizacao=datetime.now(),
                ultima_utilizacao=datetime.now(),
            )
            self.session.add(row)
            # SessionLocal usa autoflush=False. Sem este flush, duas decisões
            # iguais na mesma revisão não veem a primeira relação pendente e
            # tentam inserir duas linhas com a mesma chave única.
            self.session.flush()
        else:
            row.estado = "aprovado"
            row.confirmacoes += 1
            row.suporte += 1
            row.confianca = Decimal("1")
            row.confirmado_por_id = config.user_id
            row.ultima_utilizacao = datetime.now()
        self._edge_relation_cache[relation_key] = row

    def _learn_edge_rejection(self, material: str, edge: str, config: AssistantConfig) -> None:
        material_key = normalize_text(material)
        edge_key = normalize_text(edge)
        if not material_key or not edge_key:
            return
        relation_key = (
            material_key,
            edge_key,
            config.user_id,
            client_key(config.client),
        )
        row = self._edge_relation_cache.get(relation_key)
        if row is None:
            row = self.session.execute(
                select(ListaMaterialRelacaoOrla).where(
                    ListaMaterialRelacaoOrla.material_normalizado == material_key,
                    ListaMaterialRelacaoOrla.orla_normalizada == edge_key,
                    ListaMaterialRelacaoOrla.user_id == config.user_id,
                    ListaMaterialRelacaoOrla.cliente_chave == client_key(config.client),
                )
            ).scalar_one_or_none()
        if row is None:
            row = ListaMaterialRelacaoOrla(
                material_normalizado=material_key,
                orla_normalizada=edge_key,
                espessura_orla=extract_edge_thickness(edge),
                user_id=config.user_id,
                cliente_chave=client_key(config.client),
                origem="manual",
                estado="rejeitado",
                confianca=Decimal("1"),
                suporte=1,
                rejeicoes=1,
                confirmado_por_id=config.user_id,
                primeira_utilizacao=datetime.now(),
                ultima_utilizacao=datetime.now(),
            )
            self.session.add(row)
            self.session.flush()
        else:
            row.estado = "rejeitado"
            row.rejeicoes += 1
            row.suporte += 1
            row.confirmado_por_id = config.user_id
            row.ultima_utilizacao = datetime.now()
        self._edge_relation_cache[relation_key] = row


def read_material_table(workbook_path: Path) -> tuple[tuple[str, ...], list[MaterialRow]]:
    workbook = load_workbook(Path(workbook_path), data_only=True, read_only=True)
    try:
        if SHEET_CUTRITE not in workbook.sheetnames:
            raise ValueError(f"Folha {SHEET_CUTRITE} não encontrada.")
        sheet = workbook[SHEET_CUTRITE]
        ordered_headers = tuple(
            str(sheet.cell(2, column).value or "").strip()
            for column in range(1, sheet.max_column + 1)
            if str(sheet.cell(2, column).value or "").strip()
        )
        headers = {
            str(sheet.cell(2, column).value or "").strip(): column
            for column in range(1, sheet.max_column + 1)
            if str(sheet.cell(2, column).value or "").strip()
        }
        required = ("Descricao", "Material", "Comp", "Larg", "Qt", "Notas")
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError("Cabeçalhos em falta na LISTAGEM_CUT_RITE: " + ", ".join(missing))
        rows: list[MaterialRow] = []
        for number in range(3, sheet.max_row + 1):
            description = sheet.cell(number, headers["Descricao"]).value
            material = sheet.cell(number, headers["Material"]).value
            if description in (None, "") and material in (None, ""):
                continue
            source_value = None
            for key in ("SourceID", "ID"):
                if key in headers:
                    source_value = sheet.cell(number, headers[key]).value
                    if source_value not in (None, ""):
                        break
            source_id = str(source_value or f"ROW-{number}")
            edges = {
                name: str(sheet.cell(number, headers[name]).value or "")
                for name in EDGE_FIELDS
                if name in headers
            }
            rows.append(
                MaterialRow(
                    row_number=number,
                    source_id=source_id,
                    description=str(description or ""),
                    material=str(material or ""),
                    length=_decimal(sheet.cell(number, headers["Comp"]).value),
                    width=_decimal(sheet.cell(number, headers["Larg"]).value),
                    quantity=_decimal(sheet.cell(number, headers["Qt"]).value),
                    article=str(sheet.cell(number, headers.get("Artigo", 1)).value or ""),
                    notes=str(sheet.cell(number, headers["Notas"]).value or ""),
                    edges=edges,
                    values={
                        header: str(sheet.cell(number, headers[header]).value or "")
                        for header in ordered_headers
                    },
                )
            )
        return ordered_headers, rows
    finally:
        workbook.close()


def read_material_rows(workbook_path: Path) -> list[MaterialRow]:
    """Compatibilidade para consumidores que precisam apenas das peças."""
    return read_material_table(workbook_path)[1]


def prepare_workbook_for_assistant(
    workbook_path: Path,
    *,
    config: AssistantConfig,
    user_name: str,
) -> int:
    """Repara a camada do assistente e atribui SourceID após a macro antiga.

    A macro ``AUTOMATION`` pode recriar a tabela e deixar AA/AB vazias. Esta
    preparação só toca nas folhas técnicas e nas colunas posteriores a Z.
    """
    path = Path(workbook_path)
    if not path.is_file():
        raise ValueError(f"Excel Lista Material não encontrado:\n{path}")
    win32_client = importlib.import_module("win32com.client")
    excel = None
    workbook = None
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=False)
        if workbook.ReadOnly:
            raise RuntimeError(
                "O Excel está aberto ou bloqueado. Guarde e feche o livro antes de o analisar."
            )

        def _sheet(name: str):
            try:
                return workbook.Worksheets.Item(name)
            except Exception:
                created = workbook.Worksheets.Add(
                    After=workbook.Worksheets.Item(workbook.Worksheets.Count)
                )
                created.Name = name
                return created

        assistant = _sheet("ASSISTENTE")
        assistant_rows = (
            ("Definição", "Valor", "Origem/observação"),
            (
                "Estado catálogo placas",
                "Disponível" if config.board_catalog_available else "Modo histórico/manual",
                config.board_catalog_message,
            ),
            (
                "Puxador da obra",
                config.handle,
                "Sugestão editável; admite exceções por Artigo/RP.",
            ),
            (
                "Exceções de puxador",
                "; ".join(f"{key}={value}" for key, value in config.handle_exceptions.items()),
                "Substituições por Artigo/RP.",
            ),
            (
                "Nota para CNC_FRESAR",
                config.cnc_note,
                "Texto curto aplicado em Notas quando uma orla contém CNC_FRESAR; pode ficar vazio.",
            ),
            (
                "Lacagem formal",
                "Ativa" if config.formal_lacquering else "Desativada",
                "Tipo_Lacagem mantém-se em Z apenas por compatibilidade.",
            ),
            (
                "Módulos ativos",
                ", ".join(name for name, enabled in config.modules.items() if enabled),
                "",
            ),
            (
                "Validação MP",
                "LEGADA",
                "O novo motor placa-orla não consulta a folha MP.",
            ),
            (
                "Passo seguinte",
                "Analisar/Completar Lista Material",
                "Depois de Importar CSV IMOS e executar AUTOMATION, guarde e feche o Excel; "
                "no Martelo V3 use CUT-RITE > Analisar/Completar Lista Material.",
            ),
        )
        for row_number, values in enumerate(assistant_rows, start=1):
            for column, value in enumerate(values, start=1):
                assistant.Cells.Item(row_number, column).Value2 = value
        assistant.Range("A1:C1").Font.Bold = True
        assistant.Columns.AutoFit()

        technical_headers = {
            "SUGESTOES": (
                "Estado", "SourceID", "Folha", "Linha", "Campo", "Original",
                "Sugerido", "Motivo", "Confiança", "Decisão",
            ),
            "VALIDACAO": (
                "Estado", "SourceID", "Regra", "Material", "Orla", "Resultado", "Explicação",
            ),
            "LOG": (
                "Data", "Utilizador", "Ação", "SourceID", "Campo", "Original", "Novo", "Motivo",
            ),
            "RAW_IMOS": ("SourceID", "Origem", "Importado_em", "Dados_IMOS"),
        }
        for sheet_name, headers in technical_headers.items():
            technical = _sheet(sheet_name)
            for column, header in enumerate(headers, start=1):
                technical.Cells.Item(1, column).Value2 = header
            technical.Range(
                technical.Cells.Item(1, 1), technical.Cells.Item(1, len(headers))
            ).Font.Bold = True

        cutrite = workbook.Worksheets.Item(SHEET_CUTRITE)
        cutrite.Cells.Item(2, 27).Value2 = "SourceID"
        cutrite.Cells.Item(2, 28).Value2 = "Estado_Assistente"
        last_row = max(
            int(cutrite.Cells(cutrite.Rows.Count, 1).End(-4162).Row),
            int(cutrite.Cells(cutrite.Rows.Count, 2).End(-4162).Row),
        )
        assigned = 0
        for row_number in range(3, last_row + 1):
            if not str(cutrite.Cells.Item(row_number, 1).Value or "").strip() and not str(
                cutrite.Cells.Item(row_number, 2).Value or ""
            ).strip():
                continue
            if not str(cutrite.Cells.Item(row_number, 27).Value or "").strip():
                cutrite.Cells.Item(row_number, 27).Value2 = f"SRC-{row_number - 2:06d}"
                assigned += 1
            if not str(cutrite.Cells.Item(row_number, 28).Value or "").strip():
                cutrite.Cells.Item(row_number, 28).Value2 = "POR_ANALISAR"

        if assigned:
            log = _sheet("LOG")
            log_row = int(log.Cells(log.Rows.Count, 1).End(-4162).Row) + 1
            log.Cells.Item(log_row, 1).Value2 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log.Cells.Item(log_row, 2).Value2 = user_name
            log.Cells.Item(log_row, 3).Value2 = "Tabela preparada para análise"
            log.Cells.Item(log_row, 8).Value2 = f"{assigned} SourceID atribuídos após AUTOMATION."
        workbook.Save()
        return assigned
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()


def apply_workbook_decisions(
    workbook_path: Path,
    decisions: Iterable[AssistantDecision],
    *,
    user_name: str,
) -> int:
    """Aplica apenas decisões humanas explícitas através do próprio Excel."""
    decisions = list(decisions)
    applicable = [
        item
        for item in decisions
        if item.action in {"aceitar", "editar"}
        and item.suggestion.row_number
    ]
    if not decisions:
        return 0
    win32_client = importlib.import_module("win32com.client")
    excel = None
    workbook = None
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass
        workbook = excel.Workbooks.Open(str(Path(workbook_path).resolve()), ReadOnly=False)
        if workbook.ReadOnly:
            raise RuntimeError("O Excel está aberto ou bloqueado; feche-o antes de aplicar as decisões.")
        sheet = workbook.Worksheets.Item(SHEET_CUTRITE)
        header_map: dict[str, int] = {}
        last_col = int(sheet.UsedRange.Columns.Count)
        for column in range(1, last_col + 1):
            header = str(sheet.Cells.Item(2, column).Value or "").strip()
            if header:
                header_map[header] = column
        if "SourceID" not in header_map:
            header_map["SourceID"] = 27
            sheet.Cells.Item(2, 27).Value2 = "SourceID"
        if "Estado_Assistente" not in header_map:
            header_map["Estado_Assistente"] = 28
            sheet.Cells.Item(2, 28).Value2 = "Estado_Assistente"
        def _sheet(name: str):
            try:
                return workbook.Worksheets.Item(name)
            except Exception:
                created = workbook.Worksheets.Add(After=workbook.Worksheets.Item(workbook.Worksheets.Count))
                created.Name = name
                return created

        suggestions_sheet = _sheet("SUGESTOES")
        suggestion_headers = (
            "Estado", "SourceID", "Folha", "Linha", "Campo", "Original",
            "Sugerido", "Motivo", "Confiança", "Decisão",
        )
        if not str(suggestions_sheet.Cells.Item(1, 1).Value or "").strip():
            for column, header in enumerate(suggestion_headers, start=1):
                suggestions_sheet.Cells.Item(1, column).Value2 = header
        suggestion_row = int(suggestions_sheet.Cells(suggestions_sheet.Rows.Count, 1).End(-4162).Row) + 1

        validation = _sheet("VALIDACAO")
        validation_headers = (
            "Estado", "SourceID", "Regra", "Material", "Orla", "Resultado", "Explicação",
        )
        if not str(validation.Cells.Item(1, 1).Value or "").strip():
            for column, header in enumerate(validation_headers, start=1):
                validation.Cells.Item(1, column).Value2 = header
        validation_row = int(validation.Cells(validation.Rows.Count, 1).End(-4162).Row) + 1

        log = _sheet("LOG")
        log_headers = ("Data", "Utilizador", "Ação", "SourceID", "Campo", "Original", "Novo", "Motivo")
        if not str(log.Cells.Item(1, 1).Value or "").strip():
            for column, header in enumerate(log_headers, start=1):
                log.Cells.Item(1, column).Value2 = header
        log_row = int(log.Cells(log.Rows.Count, 1).End(-4162).Row) + 1
        applied = 0
        rows_to_delete: set[int] = set()
        applicable_ids = {id(item) for item in applicable}
        for decision in decisions:
            suggestion = decision.suggestion
            value = decision.value if decision.action == "editar" else suggestion.suggested
            if suggestion.row_number:
                sheet.Cells.Item(
                    suggestion.row_number, header_map["SourceID"]
                ).Value2 = suggestion.source_id
                sheet.Cells.Item(
                    suggestion.row_number, header_map["Estado_Assistente"]
                ).Value2 = "REVISTO"
            if id(decision) in applicable_ids and suggestion.delete_row:
                rows_to_delete.add(int(suggestion.row_number))
                applied += 1
            elif id(decision) in applicable_ids:
                column = header_map.get(suggestion.field)
                if column and suggestion.field not in {"SourceID", "Estado_Assistente"}:
                    sheet.Cells.Item(suggestion.row_number, column).Value2 = value
                    applied += 1

            suggestion_values = (
                "BLOQUEIO" if suggestion.blocking else "PROPOSTA",
                suggestion.source_id,
                SHEET_CUTRITE,
                suggestion.row_number or "",
                suggestion.field,
                suggestion.original,
                value,
                suggestion.reason,
                suggestion.confidence,
                decision.action,
            )
            for column, cell_value in enumerate(suggestion_values, start=1):
                suggestions_sheet.Cells.Item(suggestion_row, column).Value2 = cell_value
            suggestion_row += 1

            if suggestion.field in EDGE_FIELDS:
                validation_values = (
                    "BLOQUEIO" if suggestion.blocking else "REVISTO",
                    suggestion.source_id,
                    suggestion.kind,
                    "",
                    value,
                    decision.action,
                    suggestion.reason,
                )
                for column, cell_value in enumerate(validation_values, start=1):
                    validation.Cells.Item(validation_row, column).Value2 = cell_value
                validation_row += 1

            log.Cells.Item(log_row, 1).Value2 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log.Cells.Item(log_row, 2).Value2 = user_name
            log.Cells.Item(log_row, 3).Value2 = decision.action
            log.Cells.Item(log_row, 4).Value2 = suggestion.source_id
            log.Cells.Item(log_row, 5).Value2 = suggestion.field
            log.Cells.Item(log_row, 6).Value2 = suggestion.original
            log.Cells.Item(log_row, 7).Value2 = value
            log.Cells.Item(log_row, 8).Value2 = suggestion.reason
            log_row += 1
        for row_number in sorted(rows_to_delete, reverse=True):
            sheet.Rows.Item(row_number).Delete()
        workbook.Save()
        return applied
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()


def _decimal_text(value: Decimal | None) -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    return format(normalized, "f")


def _bar_group_key(row: MaterialRow, *, include_width: bool) -> tuple[object, ...]:
    """Chave operacional: diferenças em Comp/Qt/Artigo não impedem agrupamento."""
    return (
        normalize_text(row.material),
        _decimal_text(row.width) if include_width else "75",
        normalize_text(row.values.get("Veio", "")),
        tuple(normalize_text(row.edges.get(side, "")) for side in EDGE_FIELDS),
        normalize_text(row.values.get("CNC_1", "")),
        normalize_text(row.values.get("CNC_2", "")),
        normalize_text(row.values.get("+comp", "")),
        normalize_text(row.values.get("+Larg", "")),
        normalize_text(row.values.get("Esp", "")),
        normalize_text(row.values.get("Esp.Mat", "")),
        normalize_text(row.values.get("Esp.Final", "")),
        normalize_text(row.notes),
    )


def _group_change_suggestions(
    group: list[MaterialRow],
    *,
    proposed_length: Decimal,
    proposed_quantity: Decimal,
    target_width: Decimal | None,
    kind: str,
    reason: str,
    confidence: float,
) -> list[AssistantSuggestion]:
    representative = group[0]
    group_id = f"{kind}:{representative.source_id}"
    sources = ", ".join(row.source_id for row in group)
    full_reason = f"{reason} Origens agrupadas: {sources}."
    result: list[AssistantSuggestion] = []
    if representative.length != proposed_length:
        result.append(
            AssistantSuggestion(
                source_id=representative.source_id,
                row_number=representative.row_number,
                field="Comp",
                original=_decimal_text(representative.length),
                suggested=_decimal_text(proposed_length),
                reason=full_reason,
                confidence=confidence,
                kind=kind,
                group_id=group_id,
            )
        )
    if representative.quantity != proposed_quantity:
        result.append(
            AssistantSuggestion(
                source_id=representative.source_id,
                row_number=representative.row_number,
                field="Qt",
                original=_decimal_text(representative.quantity),
                suggested=_decimal_text(proposed_quantity),
                reason=full_reason,
                confidence=confidence,
                kind=f"{kind}_quantidade",
                group_id=group_id,
            )
        )
    if target_width is not None and representative.width != target_width:
        result.append(
            AssistantSuggestion(
                source_id=representative.source_id,
                row_number=representative.row_number,
                field="Larg",
                original=_decimal_text(representative.width),
                suggested=_decimal_text(target_width),
                reason=full_reason,
                confidence=0.97,
                kind=f"{kind}_largura",
                group_id=group_id,
            )
        )
    for duplicate in group[1:]:
        result.append(
            AssistantSuggestion(
                source_id=duplicate.source_id,
                row_number=duplicate.row_number,
                field="__DELETE_ROW__",
                original=duplicate.description,
                suggested=f"Remover; agrupada na linha {representative.row_number}",
                reason=full_reason,
                confidence=confidence,
                kind=f"{kind}_remover_linha",
                delete_row=True,
                group_id=group_id,
            )
        )
    return result


def _vista_vertical_suggestions(
    rows: Iterable[MaterialRow],
) -> list[AssistantSuggestion]:
    matches = [
        row for row in rows if "VISTA_VERTICAL" in normalize_text(row.description)
    ]
    grouped: dict[tuple[object, ...], list[MaterialRow]] = {}
    for row in matches:
        grouped.setdefault(_bar_group_key(row, include_width=True), []).append(row)
    result: list[AssistantSuggestion] = []
    for group in grouped.values():
        measured = [row.length for row in group if row.length is not None]
        if not measured:
            continue
        proposed_length = Decimal(math.ceil(float(max(measured) + Decimal("30"))))
        proposed_quantity = sum(
            (row.quantity or Decimal("0") for row in group), Decimal("0")
        )
        result.extend(
            _group_change_suggestions(
                group,
                proposed_length=proposed_length,
                proposed_quantity=proposed_quantity,
                target_width=None,
                kind="barra_vista_vertical",
                reason=(
                    "Vista Vertical compatível (Larg/CNC/orlas iguais): maior Comp "
                    "+ 30 mm inteiros; Qt somada para acerto em obra."
                ),
                confidence=0.90,
            )
        )
    return result


def _threshold_bar_suggestions(
    rows: Iterable[MaterialRow],
    *,
    description_key: str,
    target_width: Decimal,
    split_threshold: Decimal,
    split_length: Decimal,
    medium_threshold: Decimal,
    short_length: Decimal,
) -> list[AssistantSuggestion]:
    matches = [
        row for row in rows if description_key in normalize_text(row.description)
    ]
    compatible: dict[tuple[object, ...], list[MaterialRow]] = {}
    for row in matches:
        compatible.setdefault(_bar_group_key(row, include_width=False), []).append(row)

    result: list[AssistantSuggestion] = []
    for compatible_group in compatible.values():
        buckets: dict[str, list[MaterialRow]] = {"long": [], "medium": [], "short": []}
        for row in compatible_group:
            if row.length is None:
                continue
            if row.length >= split_threshold:
                buckets["long"].append(row)
            elif row.length >= medium_threshold:
                buckets["medium"].append(row)
            else:
                buckets["short"].append(row)
        for bucket, group in buckets.items():
            if not group:
                continue
            base_quantity = sum(
                (row.quantity or Decimal("0") for row in group), Decimal("0")
            )
            if bucket == "long":
                proposed_length = split_length
                proposed_quantity = base_quantity * Decimal("2")
                rule = (
                    f"Comp >= {_decimal_text(split_threshold)}: dividir cada peça em 2, "
                    f"usar Comp {_decimal_text(split_length)} e somar Qt."
                )
            elif bucket == "medium":
                greatest = max(row.length for row in group if row.length is not None)
                proposed_length = min(
                    split_length,
                    Decimal(math.ceil(float(greatest + Decimal("30")))),
                )
                proposed_quantity = base_quantity
                rule = (
                    f"{_decimal_text(medium_threshold)} <= Comp < "
                    f"{_decimal_text(split_threshold)}: maior Comp + 30 mm, Qt somada."
                )
            else:
                proposed_length = short_length
                proposed_quantity = base_quantity
                rule = (
                    f"Comp < {_decimal_text(medium_threshold)}: usar Comp "
                    f"{_decimal_text(short_length)} e somar Qt."
                )
            result.extend(
                _group_change_suggestions(
                    group,
                    proposed_length=proposed_length,
                    proposed_quantity=proposed_quantity,
                    target_width=target_width,
                    kind=f"barra_{description_key.lower()}_{bucket}",
                    reason=(
                        f"{description_key.replace('_', ' ').title()}: {rule} "
                        f"Larg {_decimal_text(target_width)}."
                    ),
                    confidence=0.92,
                )
            )
    return result
