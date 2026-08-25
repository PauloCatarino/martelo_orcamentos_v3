"""Exportar o catálogo de matérias-primas para Excel."""

from __future__ import annotations

from datetime import date

import openpyxl

from app.services.materias_primas_excel_export import (
    exportar_materias_primas,
    nome_do_ficheiro,
)

HOJE = date(2026, 8, 25)
COLUNAS = ["Ref LE", "Descrição", "Preço tabela", "Desc %", "Último preço", "Ativo"]
LINHAS = [
    (["PLC0001", "AGL FOL ALD. BÉTULA 19MM", "24,87 €", "20%", "23-04-2026", "Sim"], True),
    (["FER0033", "CABO ALIMENTACAO", "8,40 €", "5%", "23-07-2025", "Não"], False),
]


def test_ficheiro_leva_titulo_cabecalho_e_linhas(tmp_path) -> None:
    caminho = exportar_materias_primas(
        COLUNAS, LINHAS, tmp_path / "catalogo.xlsx", hoje=HOJE
    )
    folha = openpyxl.load_workbook(caminho).active

    assert "Matérias-Primas" in folha.cell(row=1, column=1).value
    assert "25-08-2026" in folha.cell(row=1, column=len(COLUNAS)).value
    assert [c.value for c in folha[2]] == COLUNAS
    assert folha.cell(row=3, column=1).value == "PLC0001"
    assert folha.cell(row=4, column=1).value == "FER0033"


def test_descontinuadas_saem_riscadas_como_no_ecra(tmp_path) -> None:
    caminho = exportar_materias_primas(
        COLUNAS, LINHAS, tmp_path / "catalogo.xlsx", hoje=HOJE
    )
    folha = openpyxl.load_workbook(caminho).active

    assert folha.cell(row=3, column=2).font.strike in (False, None)
    assert folha.cell(row=4, column=2).font.strike is True


def test_ficheiro_abre_pronto_a_consultar(tmp_path) -> None:
    """Cabeçalho fixo e filtros: é para ler, não para voltar a importar."""
    caminho = exportar_materias_primas(
        COLUNAS, LINHAS, tmp_path / "catalogo.xlsx", hoje=HOJE
    )
    folha = openpyxl.load_workbook(caminho).active

    assert folha.freeze_panes == "A3"
    assert folha.auto_filter.ref == "A2:F4"


def test_exporta_as_colunas_que_lhe_derem(tmp_path) -> None:
    """A exportação segue a escolha do utilizador, não uma lista fixa."""
    colunas = ["Descrição", "Fornecedor"]
    linhas = [(["AGL 19MM", "SONAE"], True)]

    caminho = exportar_materias_primas(colunas, linhas, tmp_path / "x.xlsx", hoje=HOJE)
    folha = openpyxl.load_workbook(caminho).active

    assert [c.value for c in folha[2]] == colunas
    assert folha.cell(row=3, column=2).value == "SONAE"


def test_catalogo_vazio_nao_rebenta(tmp_path) -> None:
    caminho = exportar_materias_primas(COLUNAS, [], tmp_path / "vazio.xlsx", hoje=HOJE)
    folha = openpyxl.load_workbook(caminho).active

    assert [c.value for c in folha[2]] == COLUNAS
    assert folha.max_row == 2


def test_nome_do_ficheiro_leva_a_data() -> None:
    assert nome_do_ficheiro(HOJE) == "Materias_Primas_2026-08-25.xlsx"


def test_cria_a_pasta_se_nao_existir(tmp_path) -> None:
    destino = tmp_path / "uma" / "pasta" / "nova" / "catalogo.xlsx"

    caminho = exportar_materias_primas(COLUNAS, LINHAS, destino, hoje=HOJE)

    assert caminho.exists()
