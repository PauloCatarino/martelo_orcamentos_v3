"""Tests for Lista Material IMOS Excel generation service."""

from __future__ import annotations

import base64
import json
import subprocess

import pytest
from openpyxl import Workbook

from app.db.base import Base
import app.models  # noqa: F401  (register all models on Base.metadata)
from app.models.producao import Producao
from app.repositories.system_setting_repository import SystemSettingRepository
from app.services.lista_material_imos_service import (
    TEMPLATE_FILENAME,
    ListaMaterialImosContext,
    _sheet_has_material_rows,
    _lista_material_imos_ps_script,
    execute_lista_material_workbook_macro,
    execute_lista_material_imos,
    prepare_lista_material_imos,
)


def test_prepare_lista_material_imos_constroi_contexto(session, tmp_path) -> None:
    folder = tmp_path / "processo"
    folder.mkdir()
    base = tmp_path / "base"
    base.mkdir()
    template = base / TEMPLATE_FILENAME
    template.write_text("template", encoding="utf-8")
    session.add(
        Producao(
            id=1,
            codigo_processo="26.1134_01_01_CLIENTE",
            ano="2026",
            num_enc_phc="1134",
            versao_obra="01",
            versao_plano="01",
            pasta_servidor=str(folder),
        )
    )
    SystemSettingRepository(session).upsert_setting(
        chave="pasta_base_dados_orcamento",
        valor=str(base),
        descricao="Pasta Base Dados Orcamento",
        tipo="pasta",
        grupo="Orcamentos",
    )
    session.commit()

    context = prepare_lista_material_imos(
        session,
        processo_id=1,
        nome_enc_imos="1134_01_26_CLIENTE",
        values={"RESPONSAVEL": "Paulo", "QTD": "3"},
    )

    values = json.loads(base64.b64decode(context.values_b64).decode("utf-8"))
    assert context.processo_id == 1
    assert context.folder_path == folder
    assert context.template_path == template
    assert context.output_path == folder / "Lista_Material_1134_01_26_CLIENTE.xlsm"
    assert values == {"RESPONSAVEL": "Paulo", "QTD": "3"}


def test_prepare_lista_material_imos_valida_pasta_e_template(session, tmp_path) -> None:
    session.add(
        Producao(
            id=1,
            codigo_processo="26.1134_01_01_CLIENTE",
            ano="2026",
            num_enc_phc="1134",
            versao_obra="01",
            versao_plano="01",
            pasta_servidor=str(tmp_path / "nao_existe"),
        )
    )
    session.commit()

    with pytest.raises(ValueError, match="Pasta do processo nao encontrada"):
        prepare_lista_material_imos(
            session,
            processo_id=1,
            nome_enc_imos="1134",
            values={},
        )

    folder = tmp_path / "processo"
    folder.mkdir()
    session.get(Producao, 1).pasta_servidor = str(folder)
    SystemSettingRepository(session).upsert_setting(
        chave="pasta_base_dados_orcamento",
        valor=str(tmp_path),
    )
    session.commit()

    with pytest.raises(ValueError, match="Modelo Excel nao encontrado"):
        prepare_lista_material_imos(
            session,
            processo_id=1,
            nome_enc_imos="1134",
            values={},
        )


def test_execute_lista_material_imos_invoca_powershell(monkeypatch, tmp_path) -> None:
    capturado: dict[str, object] = {}

    class _TemporaryScript:
        name = str(tmp_path / "assistente.ps1")

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def write(self, value):
            capturado["script"] = value

    def _fake_tempfile(**kwargs):
        capturado["tempfile_kwargs"] = kwargs
        return _TemporaryScript()

    def _fake_run(cmd, **kwargs):
        capturado["cmd"] = cmd
        capturado["kwargs"] = kwargs
        return subprocess.CompletedProcess(cmd, 0, stdout="", stderr="")

    monkeypatch.setattr("app.services.lista_material_imos_service.subprocess.run", _fake_run)
    monkeypatch.setattr(
        "app.services.lista_material_imos_service.tempfile.NamedTemporaryFile",
        _fake_tempfile,
    )
    monkeypatch.setattr("app.services.lista_material_imos_service.os.unlink", lambda _path: None)
    context = ListaMaterialImosContext(
        processo_id=1,
        folder_path=tmp_path,
        template_path=tmp_path / TEMPLATE_FILENAME,
        output_path=tmp_path / "Lista_Material_TESTE.xlsm",
        values_b64="e30=",
    )

    output = execute_lista_material_imos(context, timeout_seconds=12)

    cmd = capturado["cmd"]
    assert output == context.output_path
    assert cmd[:6] == [
        "powershell",
        "-NoProfile",
        "-NonInteractive",
        "-ExecutionPolicy",
        "Bypass",
        "-STA",
    ]
    assert cmd[-3:] == [str(context.template_path), str(context.output_path), "e30="]
    assert capturado["kwargs"]["timeout"] == 12
    assert capturado["tempfile_kwargs"]["encoding"] == "utf-8-sig"


def test_lista_material_imos_ps_script_tem_mapeamento_excel() -> None:
    script = _lista_material_imos_ps_script()

    assert "$ws.Range('B3').Value2 = [string]$v.RESPONSAVEL" in script
    assert "$ws.Range('P3').Value2 = [string]$v.ENC_PHC" in script
    assert "$wb.SaveAs($OutputPath, 52)" in script
    assert "$newLo.Name = 'Tabela_Cut_Rite'" in script
    assert "Get-OrCreateSheet $wb 'ASSISTENTE'" in script
    assert "Get-OrCreateSheet $wb 'RAW_IMOS'" in script
    assert "Get-OrCreateSheet $wb 'SUGESTOES'" in script
    assert "Get-OrCreateSheet $wb 'VALIDACAO'" in script
    assert "Get-OrCreateSheet $wb 'LOG'" in script
    assert "$wsCutTechnical.Range('AA2').Value2 = 'SourceID'" in script
    assert "Analisar/Completar Lista Material" in script
    assert "Substituições por Artigo/RP." in script


def test_payload_unicode_e_transportado_com_escapes_ascii(session, tmp_path) -> None:
    folder = tmp_path / "processo"
    folder.mkdir()
    template = tmp_path / TEMPLATE_FILENAME
    template.write_text("template", encoding="utf-8")
    session.add(
        Producao(
            id=1,
            codigo_processo="26.1313_01_01_JF_VIVA",
            ano="2026",
            num_enc_phc="1313",
            versao_obra="01",
            versao_plano="01",
            pasta_servidor=str(folder),
        )
    )
    SystemSettingRepository(session).upsert_setting(
        chave="pasta_base_dados_orcamento", valor=str(tmp_path)
    )
    session.commit()

    context = prepare_lista_material_imos(
        session,
        processo_id=1,
        nome_enc_imos="1313_01_26_JF_VIVA",
        values={"TEXTO": "Substituições e exceções"},
    )

    raw = base64.b64decode(context.values_b64).decode("utf-8")
    assert "Substituições" not in raw
    assert "\\u00e7" in raw
    assert json.loads(raw)["TEXTO"] == "Substituições e exceções"


def test_macro_excel_e_executada_visivel_com_argumentos(monkeypatch, tmp_path) -> None:
    path = tmp_path / "Lista_Material_TESTE.xlsm"
    path.write_bytes(b"excel")
    calls: dict[str, object] = {}

    class _Workbook:
        Name = path.name
        ReadOnly = False

        def Save(self):
            calls["saved"] = True

        def Close(self, save):
            calls["closed"] = save

    workbook = _Workbook()

    class _Workbooks:
        def Open(self, opened, ReadOnly=False):
            calls["opened"] = (opened, ReadOnly)
            return workbook

    class _Excel:
        Workbooks = _Workbooks()
        Visible = False
        DisplayAlerts = False
        AutomationSecurity = 3

        def Run(self, *args):
            calls["run"] = args

        def Quit(self):
            calls["quit"] = True

    excel = _Excel()

    class _Client:
        @staticmethod
        def DispatchEx(name):
            calls["dispatch"] = name
            return excel

    monkeypatch.setattr(
        "app.services.lista_material_imos_service.importlib.import_module",
        lambda _name: _Client,
    )

    result = execute_lista_material_workbook_macro(path, "MinhaMacro", True)

    assert result == path
    assert excel.Visible is True
    assert excel.AutomationSecurity == 1
    assert calls["run"] == (f"'{path.name}'!MinhaMacro", True)
    assert calls["saved"] is True
    assert calls["closed"] is False
    assert calls["quit"] is True


def test_validacao_dos_passos_confirma_se_a_folha_recebeu_pecas(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    ordered = workbook.active
    ordered.title = "LISTA_ORDENADA"
    ordered.cell(4, 2, "Porta Direita")
    cutrite = workbook.create_sheet("LISTAGEM_CUT_RITE")
    workbook.save(path)

    assert _sheet_has_material_rows(
        path, "LISTA_ORDENADA", start_row=4, description_column=2
    )
    assert not _sheet_has_material_rows(
        path, "LISTAGEM_CUT_RITE", start_row=3, description_column=1
    )
