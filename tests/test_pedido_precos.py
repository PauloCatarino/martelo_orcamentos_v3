"""Pedido de preços aos fornecedores: agrupamento, anexo e email."""

from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import openpyxl
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication

from app.domain.materia_prima_types import TIPO_PRECO_LIVRE, TIPO_PRECO_TABELA
from app.domain.pedido_precos import (
    COLUNA_CODIGO,
    COLUNA_DESCONTO,
    COLUNA_PRECO_ATUAL,
    COLUNA_PRECO_NOVO,
    COLUNAS_ANEXO,
    agrupar_por_fornecedor,
    assunto_do_email,
    corpo_do_email,
    materiais_a_rever,
    nome_do_anexo,
)
from app.services.pedido_precos_service import escrever_anexo
from app.ui.dialogs.pedido_precos_dialog import PedidoPrecosDialog

_app = QApplication.instance() or QApplication([])

HOJE = date(2026, 8, 25)


def _materia(**overrides):
    base = {
        "id": 1,
        "ref_le": "PLC0052",
        "descricao": "AGL TERM BEGE ARDENNE 19MM",
        "referencia_fornecedor": "413/BRILHO",
        "unidade": "M2",
        "preco_tabela": Decimal("31.20"),
        "preco_liquido": Decimal("25.58"),
        "data_ultimo_preco": date(2025, 7, 23),
        "tipo_preco": TIPO_PRECO_TABELA,
        "fornecedor": "SONAE",
        "fornecedor_id": 1,
        "ativo": True,
    }
    base.update(overrides)
    return SimpleNamespace(**base)


def _fornecedor(**overrides):
    base = {
        "id": 1,
        "nome": "SONAE",
        "email": "comercial@sonae.pt",
        "email_cc": None,
        "pessoa_contacto": None,
    }
    base.update(overrides)
    return SimpleNamespace(**base)


# ------------------------------------------------------------- agrupamento


def test_so_entram_os_precos_velhos_e_ativos() -> None:
    materias = [
        _materia(id=1),
        _materia(id=2, data_ultimo_preco=date(2026, 6, 1)),  # recente
        _materia(id=3, ativo=False),  # descontinuada
        _materia(id=4, tipo_preco=TIPO_PRECO_LIVRE, preco_liquido=None),  # preço livre
    ]

    a_rever = materiais_a_rever(materias, HOJE)

    assert [m.id for m in a_rever] == [1]


def test_agrupa_por_fornecedor_e_poe_os_sem_email_no_fim() -> None:
    materias = [
        _materia(id=1, fornecedor_id=1),
        _materia(id=2, fornecedor_id=1),
        _materia(id=3, fornecedor_id=2, fornecedor="EGGER"),
    ]
    fornecedores = [
        _fornecedor(),
        _fornecedor(id=2, nome="EGGER", email=None),
    ]

    pedidos = agrupar_por_fornecedor(materias, fornecedores, HOJE)

    assert [p.fornecedor_nome for p in pedidos] == ["SONAE", "EGGER"]
    assert pedidos[0].total == 2
    assert pedidos[0].tem_email is True
    assert pedidos[1].tem_email is False


def test_materiais_sem_fornecedor_ficam_num_grupo_proprio() -> None:
    materias = [_materia(id=1, fornecedor_id=None, fornecedor=None)]

    pedidos = agrupar_por_fornecedor(materias, [], HOJE)

    assert pedidos[0].fornecedor_nome == "(sem fornecedor)"
    assert pedidos[0].tem_email is False


def test_preco_mais_antigo_do_lote() -> None:
    materias = [
        _materia(id=1, data_ultimo_preco=date(2025, 7, 23)),
        _materia(id=2, data_ultimo_preco=date(2025, 3, 10)),
    ]

    pedido = agrupar_por_fornecedor(materias, [_fornecedor()], HOJE)[0]

    assert pedido.preco_mais_antigo == date(2025, 3, 10)


# ------------------------------------------------------------------ anexo


def test_anexo_leva_so_o_que_o_fornecedor_deve_ver(tmp_path) -> None:
    """O desconto, a margem, o preço líquido e o desperdício são nossos."""
    pedido = agrupar_por_fornecedor([_materia()], [_fornecedor()], HOJE)[0]

    caminho = escrever_anexo(pedido, tmp_path / "pedido.xlsx")
    folha = openpyxl.load_workbook(caminho).active

    cabecalhos = [celula.value for celula in folha[1]]
    assert cabecalhos == list(COLUNAS_ANEXO)
    assert "Preço líquido" not in cabecalhos
    assert "Margem" not in cabecalhos
    assert "Desperdício" not in cabecalhos
    assert "Stock" not in cabecalhos


def test_anexo_leva_o_preco_atual_e_deixa_o_novo_em_branco(tmp_path) -> None:
    pedido = agrupar_por_fornecedor([_materia()], [_fornecedor()], HOJE)[0]

    caminho = escrever_anexo(pedido, tmp_path / "pedido.xlsx")
    folha = openpyxl.load_workbook(caminho).active
    colunas = {celula.value: celula.column for celula in folha[1]}

    assert folha.cell(row=2, column=colunas[COLUNA_CODIGO]).value == "PLC0052"
    assert folha.cell(row=2, column=colunas[COLUNA_PRECO_ATUAL]).value == 31.2
    assert folha.cell(row=2, column=colunas[COLUNA_PRECO_NOVO]).value is None
    assert folha.cell(row=2, column=colunas[COLUNA_DESCONTO]).value is None


def test_anexo_bloqueia_o_codigo_e_liberta_o_que_e_para_preencher(tmp_path) -> None:
    pedido = agrupar_por_fornecedor([_materia()], [_fornecedor()], HOJE)[0]

    caminho = escrever_anexo(pedido, tmp_path / "pedido.xlsx")
    folha = openpyxl.load_workbook(caminho).active
    colunas = {celula.value: celula.column for celula in folha[1]}

    assert folha.protection.sheet is True
    assert folha.cell(row=2, column=colunas[COLUNA_CODIGO]).protection.locked is True
    assert folha.cell(row=2, column=colunas[COLUNA_PRECO_NOVO]).protection.locked is False
    assert folha.cell(row=2, column=colunas[COLUNA_DESCONTO]).protection.locked is False


def test_nome_do_anexo_leva_fornecedor_e_data() -> None:
    pedido = agrupar_por_fornecedor([_materia()], [_fornecedor()], HOJE)[0]

    assert nome_do_anexo(pedido, HOJE) == "Precos_LancaEncanto_SONAE_2026-08-25.xlsx"


def test_nome_do_anexo_aguenta_nomes_com_simbolos() -> None:
    pedido = agrupar_por_fornecedor(
        [_materia(fornecedor_id=9)], [_fornecedor(id=9, nome="B&F / Lda")], HOJE
    )[0]

    nome = nome_do_anexo(pedido, HOJE)

    assert "/" not in nome and "&" not in nome
    assert nome.endswith(".xlsx")


# ------------------------------------------------------------------ email


def test_email_diz_o_que_e_para_preencher_e_nao_expoe_o_nosso_calculo() -> None:
    pedido = agrupar_por_fornecedor([_materia()], [_fornecedor()], HOJE)[0]

    assunto = assunto_do_email(pedido)
    corpo = corpo_do_email(pedido, "anexo.xlsx", remetente="Paulo Catarino")

    assert "1 referência" in assunto
    assert COLUNA_PRECO_NOVO in corpo
    assert COLUNA_DESCONTO in corpo
    assert "Paulo Catarino" in corpo
    assert "anexo.xlsx" in corpo
    assert "líquido" not in corpo
    assert "margem" not in corpo.lower()


def test_email_trata_a_pessoa_pelo_nome_quando_ha_contacto() -> None:
    pedido = agrupar_por_fornecedor(
        [_materia()], [_fornecedor(pessoa_contacto="Ana")], HOJE
    )[0]

    assert "Bom dia, Ana" in corpo_do_email(pedido, "anexo.xlsx")


# ----------------------------------------------------------------- diálogo


def test_dialogo_marca_so_quem_tem_email() -> None:
    materias = [_materia(id=1), _materia(id=3, fornecedor_id=2, fornecedor="EGGER")]
    pedidos = agrupar_por_fornecedor(
        materias, [_fornecedor(), _fornecedor(id=2, nome="EGGER", email=None)], HOJE
    )

    dialogo = PedidoPrecosDialog(pedidos)

    assert dialogo.table.item(0, 0).checkState() == Qt.CheckState.Checked
    assert dialogo.table.item(1, 0).checkState() == Qt.CheckState.Unchecked
    assert [p.fornecedor_nome for p in dialogo.pedidos_escolhidos()] == ["SONAE"]
    assert "1 sem email" in dialogo.status_label.text()


def test_dialogo_nunca_prepara_para_quem_nao_tem_email() -> None:
    pedidos = agrupar_por_fornecedor(
        [_materia(fornecedor_id=2, fornecedor="EGGER")],
        [_fornecedor(id=2, nome="EGGER", email=None)],
        HOJE,
    )
    preparados: list = []
    dialogo = PedidoPrecosDialog(pedidos, on_preparar=lambda p: preparados.append(p) or True)

    # Mesmo marcando à força, quem não tem email fica de fora.
    dialogo.table.item(0, 0).setCheckState(Qt.CheckState.Checked)
    dialogo._preparar()

    assert preparados == []
    assert "com email preenchido" in dialogo.status_label.text()


def test_dialogo_sem_nada_a_rever() -> None:
    dialogo = PedidoPrecosDialog([])

    assert "nada a pedir" in dialogo.status_label.text()
