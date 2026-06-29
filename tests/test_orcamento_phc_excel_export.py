"""Teste do gerador do Excel no formato PHC (C2b)."""

from __future__ import annotations

from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.orcamento_phc_excel_export import gerar_excel_phc


def _items() -> list[SimpleNamespace]:
    return [
        SimpleNamespace(
            codigo="RP_01(A)",
            # Título + linha "- " + linha "* " (linha vazia ignorada).
            descricao=(
                "Movel de cozinha em termolaminado\n"
                "- Puxador TIC-TAC\n"
                "\n"
                "* Montado em obra"
            ),
            altura=Decimal("720"),
            largura=Decimal("600"),
            profundidade=Decimal("560"),
            quantidade=Decimal("2"),
            unidade="und",
            preco_unitario=Decimal("1191.62"),
        ),
        SimpleNamespace(
            codigo="RP_02",
            descricao=None,
            altura=None,
            largura=Decimal("800"),
            profundidade=None,
            quantidade=Decimal("1"),
            unidade="",
            preco_unitario=Decimal("50"),
        ),
    ]


def test_gerar_excel_phc_cria_folha_e_cabecalho(tmp_path) -> None:
    output = tmp_path / "260001_01_PHC.xlsx"
    orcamento = SimpleNamespace(num_orcamento="260001", numero_versao=1)

    resultado = gerar_excel_phc(output, orcamento=orcamento, items=_items())

    assert resultado == output
    assert output.exists()

    wb = load_workbook(output)
    assert wb.sheetnames == ["PHC"]
    ws = wb["PHC"]

    cabecalho = [cell.value for cell in ws[1]]
    assert cabecalho == [
        "RefCliente",
        "Referencia",
        "Designacao",
        "XAltura",
        "YLargura",
        "ZEspessura",
        "Qtd",
        "Und",
        "Venda",
    ]


def test_gerar_excel_phc_linha_principal_e_extra(tmp_path) -> None:
    output = tmp_path / "260001_01_PHC.xlsx"
    orcamento = SimpleNamespace(num_orcamento="260001", numero_versao=1)

    gerar_excel_phc(output, orcamento=orcamento, items=_items())

    ws = load_workbook(output)["PHC"]

    # Linha 2: linha principal do 1.º item.
    assert ws.cell(row=2, column=1).value == "RP_01(A)"
    assert ws.cell(row=2, column=2).value == "MOB"
    designacao = ws.cell(row=2, column=3).value
    assert designacao.startswith("COMP. MOB. - ")
    assert "MOVEL DE COZINHA EM TERMOLAMINADO" in designacao
    # Dimensões e quantidade como números.
    assert ws.cell(row=2, column=4).value == 720
    assert ws.cell(row=2, column=5).value == 600
    assert ws.cell(row=2, column=6).value == 560
    assert ws.cell(row=2, column=7).value == 2
    # "und" -> "un".
    assert ws.cell(row=2, column=8).value == "un"
    # Venda como TEXTO "1191,62" com number_format "@".
    venda_cell = ws.cell(row=2, column=9)
    assert venda_cell.value == "1191,62"
    assert isinstance(venda_cell.value, str)
    assert venda_cell.number_format == "@"

    # Linhas extra: só a coluna C (Designacao) preenchida.
    extra_traco = ws.cell(row=3, column=3).value
    assert extra_traco == "- Puxador TIC-TAC"
    assert ws.cell(row=3, column=1).value in (None, "")
    assert ws.cell(row=3, column=9).value in (None, "")
    extra_estrela = ws.cell(row=4, column=3).value
    assert extra_estrela == "* Montado em obra"

    # A linha vazia da descrição não gera linha extra: o 2.º item segue-se já.
    assert ws.cell(row=5, column=1).value == "RP_02"


def test_gerar_excel_phc_item_sem_descricao_e_unidade(tmp_path) -> None:
    output = tmp_path / "260001_01_PHC.xlsx"
    orcamento = SimpleNamespace(num_orcamento="260001", numero_versao=1)

    gerar_excel_phc(output, orcamento=orcamento, items=_items())

    ws = load_workbook(output)["PHC"]

    # 2.º item (linha 5): sem descrição -> só o prefixo sem traço final.
    assert ws.cell(row=5, column=3).value == "COMP. MOB. -"
    # Unidade vazia -> "un" por defeito.
    assert ws.cell(row=5, column=8).value == "un"
    # Altura/profundidade None ficam vazias; largura é número.
    assert ws.cell(row=5, column=4).value in (None, "")
    assert ws.cell(row=5, column=5).value == 800
    assert ws.cell(row=5, column=9).value == "50,00"
