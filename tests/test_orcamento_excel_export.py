"""Teste do gerador do Excel do orçamento (fase 8W.4.2)."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace
from xml.etree import ElementTree as ET
import zipfile

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

from app.domain.relatorio_totais import calcular_totais_relatorio
from app.services.orcamento_excel_export import gerar_excel_orcamento


def _rich_blocks(value) -> list[TextBlock]:
    assert isinstance(value, CellRichText)
    assert all(isinstance(part, TextBlock) for part in value)
    return list(value)


def _items() -> list[SimpleNamespace]:
    return [
        SimpleNamespace(
            ordem=1,
            codigo="A1",
            descricao="Móvel de cozinha\n- Puxador TIC-TAC\n\n* Montado",
            item="Móvel",
            altura=Decimal("720"),
            largura=Decimal("600"),
            profundidade=Decimal("560"),
            unidade="un",
            quantidade=Decimal("2"),
            preco_unitario=Decimal("100"),
            preco_total=Decimal("200"),
        ),
        SimpleNamespace(
            ordem=2,
            codigo=None,
            descricao=None,
            item="Prateleira",
            altura=None,
            largura=Decimal("800"),
            profundidade=None,
            unidade="un",
            quantidade=Decimal("1"),
            preco_unitario=Decimal("50"),
            preco_total=Decimal("50"),
        ),
    ]


def test_gerar_excel_orcamento_cria_ficheiro_com_dados(tmp_path) -> None:
    items = _items()
    totais = calcular_totais_relatorio(items)
    cliente = SimpleNamespace(
        nome="JF & Filhos, Lda",
        nome_simplex="JF VIVA",
        morada="Rua A, 1",
        email="a@b.pt",
        telefone="912345678",
        num_cliente="C123",
    )
    orcamento = SimpleNamespace(
        num_orcamento="260655",
        numero_versao=1,
        ano=2026,
        obra="Cozinha",
        ref_cliente="REF-9",
        created_at=datetime(2026, 6, 18),
    )

    saida = tmp_path / "x.xlsx"
    resultado = gerar_excel_orcamento(
        saida, cliente=cliente, orcamento=orcamento, items=items, totais=totais
    )

    assert resultado == saida
    assert saida.exists()
    assert saida.stat().st_size > 0

    with zipfile.ZipFile(saida) as xlsx:
        nomes = set(xlsx.namelist())
        ET.fromstring(xlsx.read("xl/worksheets/sheet1.xml"))
        if "xl/sharedStrings.xml" in nomes:
            ET.fromstring(xlsx.read("xl/sharedStrings.xml"))

    wb = load_workbook(saida, rich_text=True)
    ws = wb.active
    assert ws.title == "Orçamento"

    # Bloco de cabeçalho.
    assert ws["D1"].value == "Orçamento"
    assert ws["D2"].value == "Nº Orçamento: 260655_01"
    assert ws["A3"].value == "JF & Filhos, Lda"
    assert ws["D3"].value == "Data: 2026-06-18"
    assert ws["A7"].value == "Ref.: REF-9"
    assert ws["D7"].value == "Obra: Cozinha"

    # Cabeçalhos da tabela na linha 10.
    assert ws["A10"].value == "Item"
    assert ws["J10"].value == "Preço Total"

    # Primeira linha de item: texto + números reais.
    assert ws["A11"].value == "1"
    blocks = _rich_blocks(ws["C11"].value)
    assert (
        "".join(block.text for block in blocks)
        == "A1 - Móvel\nMóvel de cozinha\n  - Puxador TIC-TAC\n\n  Montado"
    )
    assert blocks[0].text == "A1 - Móvel"
    assert blocks[0].font.b is True
    assert blocks[1].text == "\nMóvel de cozinha"
    assert blocks[1].font.b is True
    assert blocks[2].text == "\n  - Puxador TIC-TAC"
    assert blocks[2].font.i is True
    assert blocks[3].text == "\n"
    assert blocks[4].text == "\n  Montado"
    assert blocks[4].font.i is True
    assert blocks[4].font.color.rgb == "FF0A5C0A"
    assert ws["C11"].alignment.wrap_text is True
    assert ws["C11"].alignment.vertical == "top"
    assert ws.column_dimensions["C"].width == 60
    assert ws["A10"].border.left.style == "thin"
    assert ws["J12"].border.bottom.style == "thin"
    assert ws["A11"].alignment.horizontal == "center"
    assert ws["G11"].alignment.horizontal == "center"
    assert ws["D11"].alignment.horizontal == "right"
    assert ws["J11"].alignment.horizontal == "right"
    assert ws["J11"].font.b is True
    assert ws["D11"].value == 720
    assert ws["H11"].value == 2
    assert ws["J11"].value == 200
    # Valor numérico (não string); openpyxl normaliza 200.0 -> 200 ao ler.
    assert isinstance(ws["J11"].value, (int, float))
    assert not isinstance(ws["J11"].value, str)

    # Segundo item usa o 'item' quando falta a descrição.
    blocks = _rich_blocks(ws["C12"].value)
    assert str(ws["C12"].value) == "Prateleira"
    assert blocks[0].text == "Prateleira"
    assert blocks[0].font.b is True

    # Totais nas colunas I/J, após 1 linha em branco (items nas linhas 11-12).
    assert ws["I14"].value == "Total Qt:"
    assert ws["J14"].value == 3
    assert ws["I15"].value == "SubTotal:"
    assert ws["J15"].value == 250.0
    assert ws["I16"].value == "IVA (23%):"
    assert ws["J16"].value == 57.5
    assert ws["I17"].value == "Total Geral:"
    assert ws["J17"].value == 307.5
