from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from app.models.lista_material_assistente import (
    ListaMaterialBarraReceita,
    ListaMaterialModulo,
    ListaMaterialObraConfig,
    ListaMaterialPlacaSnapshot,
    ListaMaterialRelacaoOrla,
)
from app.services.lista_material_assistente_service import (
    AssistantConfig,
    ListaMaterialAssistantService,
    MaterialRow,
    compact_handle,
    extract_handle,
    normalize_text,
    read_material_table,
    read_material_rows,
)
from app.services.warehouse_board_catalog import (
    BoardRecord,
    WarehouseBoardCatalogProvider,
    sync_board_snapshot,
)


def _row(**changes) -> MaterialRow:
    values = dict(
        row_number=3,
        source_id="SRC-000001",
        description="Vista Vertical",
        material="AGL MLM CARVALHO H3171 19MM",
        length=Decimal("2400"),
        width=Decimal("80"),
        quantity=Decimal("2"),
        article="RP_01",
        notes="",
        edges={
            "Orla ESQ": "PVC_1.0_H3171",
            "Orla DIR": "PVC_1.0_H3171",
            "Orla CIMA": "",
            "Orla BAIXO": "",
        },
    )
    values.update(changes)
    return MaterialRow(**values)


def test_normalizacao_e_puxador_sao_conservadores() -> None:
    assert normalize_text("Rodapé Frente") == "RODAPE_FRENTE"
    assert extract_handle("Portas com puxador J H1030; restante normal") == "puxador J H1030"
    assert extract_handle("sem indicação") == ""
    assert compact_handle("PUXADOR 'J' H1030") == "Pux 'J' H1030"


def test_perfil_respeita_precedencia_e_jf_viva_desativa_lacagem(session) -> None:
    session.add_all(
        [
            ListaMaterialModulo(user_id=0, cliente_chave="", modulo="cnc_fresar", ativo=False),
            ListaMaterialModulo(user_id=7, cliente_chave="", modulo="cnc_fresar", ativo=True),
            ListaMaterialModulo(user_id=7, cliente_chave="JF_VIVA", modulo="cnc_fresar", ativo=False),
        ]
    )
    session.commit()

    config = ListaMaterialAssistantService(session).resolve_config(
        user_id=7, client="JF_VIVA", production_description="Puxador TIC-TAC"
    )

    assert config.modules["cnc_fresar"] is False
    assert config.modules["lacagem_formal"] is False
    assert config.formal_lacquering is False
    assert "TIC-TAC" in config.handle.upper()
    assert config.board_catalog_available is False


def test_aprendizagem_repetida_atualiza_relacao_com_autoflush_desativado(session) -> None:
    session.autoflush = False
    service = ListaMaterialAssistantService(session)
    config = AssistantConfig(user_id=2, client="JF_VIVA")

    service._learn_edge_relation(
        "AGL_MLM_LINHO_CANCUN_19MM", "PVC_0.4_LINHO", config
    )
    service._learn_edge_relation(
        "AGL_MLM_LINHO_CANCUN_19MM", "PVC_0.4_LINHO", config
    )
    session.commit()

    relations = session.query(ListaMaterialRelacaoOrla).all()
    assert len(relations) == 1
    assert relations[0].suporte == 2
    assert relations[0].confirmacoes == 2


def test_analise_da_obra_reutiliza_snapshot_escolhido_na_criacao(session) -> None:
    session.add(
        ListaMaterialObraConfig(
            producao_id=1313,
            user_id=7,
            cliente_chave="JF_VIVA",
            workbook_path="Lista_Material_1313.xlsm",
            puxador_obra="Puxador J H1030",
            puxadores_excecoes_json=(
                '{"RP_02": "TIC-TAC", "CNC_FRESAR": "CNC RECORTE L"}'
            ),
            modulos_json='{"puxadores": true}',
            configuracao_json=(
                '{"modules": {"puxadores": true, "cnc_fresar": false}, '
                '"formal_lacquering": false}'
            ),
        )
    )
    session.commit()

    config = ListaMaterialAssistantService(session).resolve_work_config(
        production_id=1313,
        user_id=7,
        client="JF_VIVA",
        production_description="Outro puxador que não deve substituir o snapshot",
    )

    assert config.handle == "Puxador J H1030"
    assert config.handle_exceptions == {"RP_02": "TIC-TAC"}
    assert config.cnc_note == "CNC RECORTE L"
    assert config.modules["cnc_fresar"] is False
    assert config.modules["puxadores"] is True


def test_vista_vertical_sugere_maior_comprimento_mais_30(session) -> None:
    rows = [_row(), _row(row_number=4, source_id="SRC-000002", length=Decimal("2450"))]
    config = AssistantConfig(user_id=7, client="JF_VIVA")

    suggestions = ListaMaterialAssistantService(session).analyze_rows(rows, config=config)

    vista = [item for item in suggestions if item.kind == "barra_vista_vertical"]
    assert len(vista) == 1
    assert vista[0].suggested == "2480"
    assert "SRC-000001" in vista[0].reason
    assert "SRC-000002" in vista[0].reason


def test_remate_e_rodape_propõem_largura_75(session) -> None:
    rows = [
        _row(description="Remate Teto", width=Decimal("73")),
        _row(row_number=4, source_id="SRC-2", description="Rodape Frente", width=Decimal("70")),
    ]
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        rows, config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    widths = [item for item in suggestions if item.field == "Larg"]
    assert [item.suggested for item in widths] == ["75", "75"]


def test_remate_teto_lacar_e_mostrado_como_conflito_jf_viva(session) -> None:
    row = _row(
        description="Remate Teto",
        edges={"Orla ESQ": "PVC_1.0_LACAR", "Orla DIR": "", "Orla CIMA": "", "Orla BAIXO": ""},
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    conflict = next(item for item in suggestions if item.kind == "remate_teto_lacagem")
    assert conflict.suggested == ""
    assert "não lacar" in conflict.reason


def test_comprimento_remate_usa_regra_operacional_confirmada(session) -> None:
    row = _row(description="Remate Teto", length=Decimal("2450"))

    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    length = next(item for item in suggestions if item.field == "Comp")
    assert length.suggested == "2480"
    assert "maior Comp + 30 mm" in length.reason


def test_cnc_fresar_separa_operacao_e_orla(session) -> None:
    row = _row(
        edges={
            "Orla ESQ": "CNC_FRESAR",
            "Orla DIR": "PVC_1.0_H3171",
            "Orla CIMA": "PVC_1.0_H3171",
            "Orla BAIXO": "",
        }
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    edge = next(item for item in suggestions if item.kind == "cnc_fresar")
    note = next(item for item in suggestions if item.kind == "notas_assistente")
    assert edge.suggested == "PVC_1.0_H3171"
    assert edge.blocking is False
    assert note.suggested == "CNC_FRESAR"


def test_cnc_fresar_usa_texto_de_notas_configurado_na_obra(session) -> None:
    row = _row(
        notes="CNC_FRESAR",
        edges={
            "Orla ESQ": "CNC_FRESAR",
            "Orla DIR": "PVC_1.0_H3171",
            "Orla CIMA": "PVC_1.0_H3171",
            "Orla BAIXO": "",
        },
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row],
        config=AssistantConfig(
            user_id=7,
            client="JF_VIVA",
            cnc_note="CNC RECORTE L",
        ),
    )

    note = next(item for item in suggestions if item.field == "Notas")
    assert note.suggested == "CNC RECORTE L"
    assert "configuração da obra" in note.reason


def test_notas_porta_juntam_lacagem_e_puxador_curto_com_excecao(session) -> None:
    row = _row(
        description="Porta Esquerda",
        material="MDF_MR_MLM_BRANCO_B3002/MA_19MM",
        article="RP_01",
        notes="PISO 1",
        edges={"Orla ESQ": "PVC_1.0_LACAR", "Orla DIR": "", "Orla CIMA": "", "Orla BAIXO": ""},
    )
    config = AssistantConfig(
        user_id=7,
        client="JF_VIVA",
        handle="TIC-TAC",
        handle_exceptions={"RP_01": "Puxador J H1030"},
    )

    suggestions = ListaMaterialAssistantService(session).analyze_rows([row], config=config)
    note = next(item for item in suggestions if item.kind == "notas_assistente")

    assert note.suggested == "PISO 1; Lacar 1 Face + Pux J H1030"
    assert "orla LACAR" in note.reason


def test_puxador_nao_e_aplicado_a_lateral_costa_teto_ou_fundo(session) -> None:
    rows = [
        _row(description="Lateral Direita"),
        _row(row_number=4, source_id="SRC-2", description="Costa"),
        _row(row_number=5, source_id="SRC-3", description="Teto"),
        _row(row_number=6, source_id="SRC-4", description="Fundo"),
        _row(row_number=7, source_id="SRC-5", description="Porta Direita"),
    ]
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        rows,
        config=AssistantConfig(
            user_id=7, client="JF_VIVA", handle="PUXADOR 'J' H1030"
        ),
    )
    notes = [item for item in suggestions if item.kind == "notas_assistente"]
    assert [(item.row_number, item.suggested) for item in notes] == [
        (7, "Pux 'J' H1030")
    ]


def test_puxador_antigo_e_removido_de_pecas_nao_elegiveis(session) -> None:
    rows = [
        _row(description="Lateral Direita", notes="PUXADOR 'J' H1030"),
        _row(
            row_number=4,
            source_id="SRC-2",
            description="Costa",
            notes="CNC_FRESAR; PUXADOR 'J' H1030",
        ),
        _row(
            row_number=5,
            source_id="SRC-3",
            description="Porta Direita",
            notes="PUXADOR 'J' H1030",
        ),
    ]
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        rows,
        config=AssistantConfig(
            user_id=7, client="JF_VIVA", handle="PUXADOR 'J' H1030"
        ),
    )
    notes = {
        item.row_number: item
        for item in suggestions
        if item.kind == "notas_assistente"
    }
    assert notes[3].suggested == ""
    assert notes[3].allow_blank is True
    assert notes[4].suggested == "CNC_FRESAR"
    assert notes[5].suggested == "Pux 'J' H1030"


def test_frente_gaveta_so_recebe_puxador_quando_tem_lacagem(session) -> None:
    plain = _row(description="Frente de Gaveta")
    lacquered = _row(
        row_number=4,
        source_id="SRC-2",
        description="Frente de Gaveta",
        edges={"Orla ESQ": "PVC_1.0_LACAR", "Orla DIR": "", "Orla CIMA": "", "Orla BAIXO": ""},
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [plain, lacquered],
        config=AssistantConfig(user_id=7, client="JF_VIVA", handle="Puxador J H1030"),
    )
    notes = [item for item in suggestions if item.kind == "notas_assistente"]
    assert len(notes) == 1
    assert notes[0].row_number == 4
    assert notes[0].suggested == "Lacar 1 Face + Pux J H1030"


def test_teto_fundo_limpam_orlas_nao_aplicaveis_e_preservam_cnc_nas_notas(session) -> None:
    row = _row(
        description="Teto",
        edges={
            "Orla ESQ": "PVC_1.0_LINHO",
            "Orla DIR": "PVC_0.4_LINHO",
            "Orla CIMA": "CNC_FRESAR",
            "Orla BAIXO": "PVC_0.4_LINHO",
        },
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    cleared = [item for item in suggestions if item.field in {"Orla DIR", "Orla CIMA", "Orla BAIXO"}]
    assert {item.field for item in cleared} == {"Orla DIR", "Orla CIMA", "Orla BAIXO"}
    assert all(item.suggested == "" and item.allow_blank for item in cleared)
    note = next(item for item in suggestions if item.field == "Notas")
    assert note.suggested == "CNC_FRESAR"


def test_maleiro_cnc_usa_orlas_dir_cima_baixo_e_ignora_esq(session) -> None:
    row = _row(
        description="Maleiro",
        edges={
            "Orla ESQ": "PVC_1.0_LINHO",
            "Orla DIR": "PVC_0.4_LINHO",
            "Orla CIMA": "CNC_FRESAR",
            "Orla BAIXO": "PVC_0.4_LINHO",
        },
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    edge = next(item for item in suggestions if item.kind == "cnc_fresar")
    assert edge.suggested == "PVC_0.4_LINHO"
    assert edge.blocking is False


def test_laterais_cnc_limpam_apenas_lado_assinalado_e_notas(session) -> None:
    for row_number, description in enumerate(
        ("Lateral Esquerda", "Lateral Direita"), start=3
    ):
        row = _row(
            row_number=row_number,
            source_id=f"SRC-L{row_number}",
            description=description,
            notes="CNC RECORTE L",
            edges={
                "Orla ESQ": "PVC_1.0_LINHO",
                "Orla DIR": "CNC_FRESAR",
                "Orla CIMA": "PVC_0.4_LINHO",
                "Orla BAIXO": "PVC_0.4_LINHO",
            },
        )
        suggestions = ListaMaterialAssistantService(session).analyze_rows(
            [row],
            config=AssistantConfig(
                user_id=7,
                client="JF_VIVA",
                cnc_note="CNC RECORTE L",
            ),
        )

        edge_suggestions = [
            item for item in suggestions if item.field.startswith("Orla ")
        ]
        assert len(edge_suggestions) == 1
        assert edge_suggestions[0].field == "Orla DIR"
        assert edge_suggestions[0].suggested == ""
        assert edge_suggestions[0].allow_blank is True
        assert edge_suggestions[0].blocking is False

        note = next(item for item in suggestions if item.field == "Notas")
        assert note.suggested == ""
        assert note.allow_blank is True


def test_lateral_cnc_com_notas_vazias_nao_cria_nota(session) -> None:
    row = _row(
        description="Lateral Esquerda",
        notes="",
        edges={
            "Orla ESQ": "PVC_1.0_LINHO",
            "Orla DIR": "CNC_FRESAR",
            "Orla CIMA": "PVC_0.4_LINHO",
            "Orla BAIXO": "PVC_0.4_LINHO",
        },
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row],
        config=AssistantConfig(
            user_id=7,
            client="JF_VIVA",
            cnc_note="CNC RECORTE L",
        ),
    )

    assert not any(item.field == "Notas" for item in suggestions)
    edge = next(item for item in suggestions if item.field == "Orla DIR")
    assert edge.suggested == ""


def test_remate_teto_b3002_nasce_com_nota_nao_lacar(session) -> None:
    row = _row(
        description="Remate Teto",
        material="MDF_MR_MLM_BRANCO_B3002/MA_19MM",
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    note = next(item for item in suggestions if item.field == "Notas")
    assert note.suggested == "Não Lacar"


def test_vista_vertical_agrupa_qt_comp_e_remove_duplicados(session) -> None:
    rows = [
        _row(quantity=Decimal("2"), length=Decimal("2400"), values={"CNC_1": "A", "CNC_2": ""}),
        _row(row_number=4, source_id="SRC-2", quantity=Decimal("3"), length=Decimal("2450"), values={"CNC_1": "A", "CNC_2": ""}),
        _row(row_number=5, source_id="SRC-3", quantity=Decimal("1"), length=Decimal("2500"), values={"CNC_1": "B", "CNC_2": ""}),
    ]
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        rows, config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    first_group = [item for item in suggestions if item.group_id.endswith("SRC-000001")]
    assert next(item for item in first_group if item.field == "Comp").suggested == "2480"
    assert next(item for item in first_group if item.field == "Qt").suggested == "5"
    removed = next(item for item in first_group if item.delete_row)
    assert removed.row_number == 4
    assert all(item.row_number != 5 or not item.delete_row for item in suggestions)


def test_remate_e_rodape_aplicam_patamares_e_divisao(session) -> None:
    remates = [
        _row(description="Remate Teto", length=Decimal("3580"), quantity=Decimal("1"), width=Decimal("73")),
        _row(row_number=4, source_id="SRC-R2", description="Remate Teto", length=Decimal("2961"), quantity=Decimal("1"), width=Decimal("73")),
        _row(row_number=5, source_id="SRC-R3", description="Remate Teto", length=Decimal("2463"), quantity=Decimal("2"), width=Decimal("73")),
        _row(row_number=6, source_id="SRC-R4", description="Remate Teto", length=Decimal("1758"), quantity=Decimal("1"), width=Decimal("73")),
    ]
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        remates, config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    long_group = [item for item in suggestions if "remate_teto_long" in item.kind]
    assert next(item for item in long_group if item.field == "Comp").suggested == "2780"
    assert next(item for item in long_group if item.field == "Qt").suggested == "4"
    assert any(item.delete_row and item.row_number == 4 for item in long_group)
    assert next(item for item in suggestions if item.row_number == 5 and item.field == "Comp").suggested == "2493"
    assert next(item for item in suggestions if item.row_number == 6 and item.field == "Comp").suggested == "2050"

    rodape = _row(description="Rodape Frente", length=Decimal("3580"), quantity=Decimal("2"), width=Decimal("70"))
    rodape_suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [rodape], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    assert next(item for item in rodape_suggestions if item.field == "Comp").suggested == "2830"
    assert next(item for item in rodape_suggestions if item.field == "Qt").suggested == "4"


def test_cnc_ambiguo_bloqueia(session) -> None:
    row = _row(
        edges={
            "Orla ESQ": "CNC_FRESAR",
            "Orla DIR": "PVC_1.0_A",
            "Orla CIMA": "PVC_1.0_B",
            "Orla BAIXO": "",
        }
    )
    suggestions = ListaMaterialAssistantService(session).analyze_rows(
        [row], config=AssistantConfig(user_id=7, client="JF_VIVA")
    )
    edge = next(item for item in suggestions if item.kind == "cnc_fresar")
    assert edge.suggested == ""
    assert edge.blocking is True


def test_leitura_excel_preserva_contrato_e_source_id(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LISTAGEM_CUT_RITE"
    headers = [
        "Descricao", "Material", "Comp", "Larg", "Qt", "Veio", "Orla", "Cliente",
        "Ref_Cliente", "Processo", "Artigo", "Notas", "Esp", "Grafico Orlas",
        "Orla ESQ", "Orla DIR", "Orla CIMA", "Orla BAIXO", "ID", "CNC_1", "CNC_2",
        "+comp", "+Larg", "Esp.Mat", "Esp.Final", "Tipo_Lacagem", "SourceID",
    ]
    for column, value in enumerate(headers, 1):
        sheet.cell(2, column, value)
    sheet.cell(3, 1, "Costa")
    sheet.cell(3, 2, "AGL_TESTE_19MM")
    sheet.cell(3, 3, 2000)
    sheet.cell(3, 4, 600)
    sheet.cell(3, 5, 2)
    sheet.cell(3, 12, "Nota local")
    sheet.cell(3, 19, 1)
    sheet.cell(3, 27, "SRC-X")
    workbook.save(path)

    rows = read_material_rows(path)

    assert len(rows) == 1
    assert rows[0].source_id == "SRC-X"
    assert rows[0].notes == "Nota local"
    assert rows[0].length == Decimal("2000")
    columns, rows_with_context = read_material_table(path)
    assert columns[:5] == ("Descricao", "Material", "Comp", "Larg", "Qt")
    assert rows_with_context[0].values["Notas"] == "Nota local"
    assert rows_with_context[0].values["SourceID"] == "SRC-X"


def test_motor_nao_importa_catalogo_orcamentacao() -> None:
    import inspect
    import app.services.lista_material_assistente_service as module

    source = inspect.getsource(module)
    assert "DefMateriaPrima" not in source
    assert "def_materias_primas" in source  # apenas na regra explícita da docstring


def test_snapshot_homag_preserva_unidade_e_orientacao(session) -> None:
    provider = WarehouseBoardCatalogProvider(
        lambda: [
            BoardRecord(
                external_id="X230226163226",
                code="AGL_MLM_CARVALHO_H3165/ST12_19MM",
                description="AGL MLM CARVALHO H3165/ST12 2800X2070X19",
                length=Decimal("2800"),
                width=Decimal("2070"),
                thickness=Decimal("19"),
                available=Decimal("4"),
            )
        ]
    )

    assert sync_board_snapshot(session, provider) == 1
    row = session.query(ListaMaterialPlacaSnapshot).one()
    assert row.comprimento == Decimal("2800")
    assert row.largura == Decimal("2070")
    assert row.unidade == "mm"
    assert sync_board_snapshot(session, provider) == 0


def test_identificacao_placa_por_codigo_e_espessura(session) -> None:
    provider = WarehouseBoardCatalogProvider(
        lambda: [
            BoardRecord(
                external_id="X1",
                code="AGL_MLM_CARVALHO_H3165_ST12_19MM",
                description="AGL MLM CARVALHO H3165/ST12 2800X2070X19",
                length=Decimal("2800"),
                width=Decimal("2070"),
                thickness=Decimal("19"),
            )
        ]
    )
    service = ListaMaterialAssistantService(session, board_catalog=provider)

    match = service.identify_board(
        "AGL MLM CARVALHO H3165 ST12 19MM",
        thickness=Decimal("19"),
        config=AssistantConfig(user_id=7, client="JF_VIVA"),
    )

    assert match.board is not None
    assert match.board.external_id == "X1"
    assert match.confidence >= 0.9
