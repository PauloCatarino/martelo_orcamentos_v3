"""Teste do gerador do Excel Resumo de Custos (fase 8W.4.3)."""

from __future__ import annotations

from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from app.services.resumo_custos_excel_export import (
    COLUNAS_RESUMO_GERAL,
    construir_linhas_geral,
    gerar_excel_resumo_custos,
)

PLACAS_HEADERS = [
    "ref_le",
    "descricao_no_orcamento",
    "pliq",
    "und",
    "desp",
    "comp_mp",
    "larg_mp",
    "esp_mp",
    "qt_placas_utilizadas",
    "area_placa",
    "m2_consumidos",
    "custo_mp_total",
    "custo_placas_utilizadas",
    "nao_stock",
]
ORLAS_HEADERS = [
    "ref_orla",
    "espessura_orla",
    "largura_orla",
    "ml_total",
    "custo_total",
]
FERRAGENS_HEADERS = [
    "ref_le",
    "descricao_no_orcamento",
    "pliq",
    "und",
    "desp",
    "comp_mp",
    "larg_mp",
    "esp_mp",
    "qt_total",
    "spp_ml_und",
    "custo_mp_und",
    "custo_mp_total",
]
MAQUINAS_HEADERS = [
    "Operação",
    "Custo Total (€)",
    "ML Corte",
    "ML Orlado",
    "Nº Peças",
]
MARGENS_HEADERS = ["Tipo", "Percentagem (%)", "Valor (€)"]

FOLHAS = {
    "Resumo Placas": PLACAS_HEADERS,
    "Resumo Orlas": ORLAS_HEADERS,
    "Resumo Ferragens": FERRAGENS_HEADERS,
    "Resumo Maquinas_MO": MAQUINAS_HEADERS,
    "Resumo Margens": MARGENS_HEADERS,
}


def _criar_modelo(path, folhas=FOLHAS) -> None:
    workbook = Workbook()
    primeira = True
    for nome, headers in folhas.items():
        if primeira:
            ws = workbook.active
            ws.title = nome
            primeira = False
        else:
            ws = workbook.create_sheet(nome)
        ws.append(headers)
        ws.append(["EXEMPLO"] * len(headers))
    workbook.save(path)


def _resumo() -> SimpleNamespace:
    return SimpleNamespace(
        placas=[
            SimpleNamespace(
                ref_le="PL001",
                descricao_no_orcamento="Melamina branca",
                pliq=Decimal("12.50"),
                unidade="M2",
                desp=Decimal("10"),
                comp_mp=Decimal("2800"),
                larg_mp=Decimal("2070"),
                esp_mp=Decimal("19"),
                qt_placas=2,
                area_placa=Decimal("5.796"),
                m2_consumidos=Decimal("7.654"),
                custo_mp_total=Decimal("95.675"),
                custo_placa_inteira=Decimal("144.90"),
                nao_stock=True,
            ),
            SimpleNamespace(
                ref_le="PL002",
                descricao_no_orcamento="MDF cru",
                pliq=Decimal("9.75"),
                unidade="M2",
                desp=Decimal("5"),
                comp_mp=Decimal("2440"),
                larg_mp=Decimal("1220"),
                esp_mp=Decimal("16"),
                qt_placas=1,
                area_placa=Decimal("2.977"),
                m2_consumidos=Decimal("1.25"),
                custo_mp_total=Decimal("12.1875"),
                custo_placa_inteira=Decimal("29.025"),
                nao_stock=False,
            ),
        ],
        orlas=[
            SimpleNamespace(
                ref_orla="ORL04",
                espessura=Decimal("0.4"),
                largura=Decimal("23"),
                ml_total=Decimal("12.345"),
                custo_total=Decimal("6.17"),
            )
        ],
        ferragens=[
            SimpleNamespace(
                ref_le="FER001",
                descricao_no_orcamento="Dobradiça",
                pliq=Decimal("1.20"),
                unidade="UN",
                desp=Decimal("0"),
                qt_total=Decimal("3"),
                ml=Decimal("0"),
                custo_total=Decimal("12"),
            )
        ],
        maquinas=[
            SimpleNamespace(
                centro="Seccionadora (Corte)",
                custo_total=Decimal("35.5"),
                ml_corte=Decimal("21.25"),
                ml_orlado=Decimal("0"),
                num_pecas=Decimal("8"),
            )
        ],
        distribuicao=SimpleNamespace(
            categorias=[
                SimpleNamespace(nome="Placas", pct=Decimal("40"), euros=Decimal("200")),
                SimpleNamespace(nome="Margens", pct=Decimal("60"), euros=Decimal("300")),
            ],
            total_venda=Decimal("500"),
        ),
    )


def _headers(ws, total: int) -> list:
    return [ws.cell(row=1, column=col).value for col in range(1, total + 1)]


def test_gerar_excel_resumo_custos_copia_modelo_e_preenche_folhas(tmp_path) -> None:
    modelo = tmp_path / "modelo.xlsx"
    saida = tmp_path / "saida.xlsx"
    _criar_modelo(modelo)

    resultado = gerar_excel_resumo_custos(saida, modelo, resumo=_resumo())

    assert resultado == saida
    assert saida.exists()
    assert saida.stat().st_size > 0

    workbook = load_workbook(saida)
    for nome, headers in FOLHAS.items():
        assert nome in workbook.sheetnames
        assert _headers(workbook[nome], len(headers)) == headers

    ws_placas = workbook["Resumo Placas"]
    assert ws_placas["A2"].value == "PL001"
    assert ws_placas["I2"].value == 2
    assert ws_placas["N2"].value == 1
    assert ws_placas["N3"].value == 0
    assert isinstance(ws_placas["L2"].value, (int, float))
    assert not isinstance(ws_placas["L2"].value, str)
    assert ws_placas["L2"].number_format == "#,##0.00 €"

    ws_ferragens = workbook["Resumo Ferragens"]
    assert ws_ferragens["K2"].value == 4

    ws_margens = workbook["Resumo Margens"]
    assert ws_margens["A2"].value == "Placas"
    assert ws_margens["A3"].value == "Margens"
    assert ws_margens["A4"].value == "Total (Venda)"
    assert ws_margens["B4"].value is None
    assert ws_margens["C4"].value == 500


def test_gerar_excel_resumo_custos_ignora_folha_em_falta(tmp_path) -> None:
    modelo = tmp_path / "modelo_sem_orlas.xlsx"
    saida = tmp_path / "saida_sem_orlas.xlsx"
    _criar_modelo(modelo, folhas={"Resumo Placas": PLACAS_HEADERS})

    gerar_excel_resumo_custos(saida, modelo, resumo=_resumo())

    workbook = load_workbook(saida)
    assert workbook.sheetnames == ["Resumo Placas"]
    assert workbook["Resumo Placas"]["A2"].value == "PL001"


def test_construir_linhas_geral_multiplica_por_item_qt() -> None:
    linhas = [
        SimpleNamespace(
            id=1,
            orcamento_item_id=10,
            tipo_linha="PECA",
            descricao="Lateral",
            quantidade=Decimal("2"),
            custo_mp=Decimal("5"),
            ml_orla_fina=Decimal("3"),
            area_m2=Decimal("1"),
            excluir_mp=True,
            ativo=True,
        ),
        SimpleNamespace(
            id=2,
            orcamento_item_id=10,
            tipo_linha="PECA",
            descricao="Inativa",
            quantidade=Decimal("1"),
            custo_mp=Decimal("999"),
            ml_orla_fina=Decimal("999"),
            area_m2=Decimal("999"),
            excluir_mp=False,
            ativo=False,
        ),
    ]

    resultado = construir_linhas_geral(linhas, {10: Decimal("3")})

    assert len(resultado) == 1
    linha = resultado[0]
    assert linha["qt_total"] == Decimal("6")
    assert linha["custo_mp"] == Decimal("15")
    assert linha["ml_orla_fina"] == Decimal("9")
    assert linha["area_m2_und"] == Decimal("1")
    assert linha["excluir_mp"] == 1


def test_gerar_excel_resumo_custos_preenche_resumo_geral(tmp_path) -> None:
    modelo = tmp_path / "modelo_com_geral.xlsx"
    saida = tmp_path / "saida_com_geral.xlsx"
    folhas = {
        "Resumo Geral": [
            f"antigo_{indice}"
            for indice in range(len(COLUNAS_RESUMO_GERAL) + 2)
        ],
        **FOLHAS,
    }
    _criar_modelo(modelo, folhas=folhas)
    linhas_geral = construir_linhas_geral(
        [
            SimpleNamespace(
                id=1,
                orcamento_item_id=10,
                tipo_linha="PECA",
                descricao="Lateral",
                quantidade=Decimal("2"),
                custo_mp=Decimal("5"),
                ml_orla_fina=Decimal("3"),
                area_m2=Decimal("1"),
                excluir_mp=True,
                ativo=True,
            )
        ],
        {10: Decimal("3")},
    )

    gerar_excel_resumo_custos(
        saida,
        modelo,
        resumo=_resumo(),
        linhas_geral=linhas_geral,
    )

    workbook = load_workbook(saida)
    ws = workbook["Resumo Geral"]
    assert _headers(ws, len(COLUNAS_RESUMO_GERAL)) == COLUNAS_RESUMO_GERAL
    assert ws.max_column == len(COLUNAS_RESUMO_GERAL)

    qt_total_col = COLUNAS_RESUMO_GERAL.index("qt_total") + 1
    custo_mp_col = COLUNAS_RESUMO_GERAL.index("custo_mp") + 1
    assert ws.cell(row=2, column=qt_total_col).value == 6
    assert ws.cell(row=2, column=custo_mp_col).value == 15
    assert isinstance(ws.cell(row=2, column=custo_mp_col).value, (int, float))
    assert not isinstance(ws.cell(row=2, column=custo_mp_col).value, str)
