"""Ler a resposta do fornecedor e aplicar só o que for aprovado."""

from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import openpyxl
import pytest
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication

from app.domain.resposta_fornecedor import (
    ESTADO_ANOMALIA,
    ESTADO_ATUALIZA,
    ESTADO_DESCONHECIDO,
    ESTADO_DESCONTINUADO,
    ESTADO_SEM_ALTERACAO,
    ESTADO_SEM_RESPOSTA,
    diz_descontinuado,
    ler_respostas,
    mapear_colunas,
    to_decimal,
)
from app.services.def_materia_prima_service import (
    CriarDefMateriaPrimaData,
    DefMateriaPrimaService,
)
from app.services.resposta_fornecedor_service import RespostaFornecedorService, ler_folha
from app.ui.dialogs.resposta_fornecedor_dialog import RespostaFornecedorDialog

_app = QApplication.instance() or QApplication([])

HOJE = date(2026, 8, 25)
CABECALHOS = [
    "Código",
    "Ref. fornecedor",
    "Designação",
    "Und",
    "Preço tabela atual",
    "Preço tabela atualizado",
    "Desconto %",
    "Nova referência",
    "Nova designação",
    "Observações",
]


def _catalogo(**overrides):
    base = {
        "id": 1,
        "ref_le": "PLC0052",
        "descricao": "AGL TERM BEGE ARDENNE 19MM",
        "preco_tabela": Decimal("30.00"),
        "desconto": Decimal("18"),
        "margem": None,
        "referencia_fornecedor": "413/BRILHO",
    }
    base.update(overrides)
    return {base["ref_le"].upper(): SimpleNamespace(**base)}


def _linha(**valores):
    ordem = [
        "codigo", "ref_fornecedor", "designacao", "unidade", "preco_atual",
        "preco_novo", "desconto", "nova_ref", "nova_designacao", "observacoes",
    ]
    base = dict.fromkeys(ordem)
    base["codigo"] = "PLC0052"
    base.update(valores)
    return tuple(base[chave] for chave in ordem)


# ------------------------------------------------------------------ leitura


def test_le_um_preco_novo_e_calcula_a_variacao() -> None:
    propostas = ler_respostas(CABECALHOS, [_linha(preco_novo=31.2)], _catalogo())

    assert len(propostas) == 1
    assert propostas[0].estado == ESTADO_ATUALIZA
    assert propostas[0].preco_novo == Decimal("31.2")
    assert round(propostas[0].variacao, 1) == Decimal("4.0")
    assert propostas[0].sugerido is True


def test_variacao_grande_fica_por_confirmar() -> None:
    propostas = ler_respostas(CABECALHOS, [_linha(preco_novo=48.9)], _catalogo())

    assert propostas[0].estado == ESTADO_ANOMALIA
    assert propostas[0].aplicavel is True
    # Aplicável, mas não vem marcada: alguém tem de olhar.
    assert propostas[0].sugerido is False


def test_descontinuado_propoe_desativar() -> None:
    propostas = ler_respostas(
        CABECALHOS, [_linha(observacoes="Artigo descontinuado")], _catalogo()
    )

    assert propostas[0].estado == ESTADO_DESCONTINUADO
    assert propostas[0].sugerido is False


def test_codigo_desconhecido_nao_e_aplicavel() -> None:
    propostas = ler_respostas(CABECALHOS, [_linha(codigo="XPTO999", preco_novo=10)], _catalogo())

    assert propostas[0].estado == ESTADO_DESCONHECIDO
    assert propostas[0].aplicavel is False


def test_linha_por_preencher_e_preco_igual() -> None:
    propostas = ler_respostas(
        CABECALHOS,
        [_linha(), _linha(preco_novo=30.00)],
        _catalogo(),
    )

    assert propostas[0].estado == ESTADO_SEM_RESPOSTA
    assert propostas[1].estado == ESTADO_SEM_ALTERACAO
    assert not any(proposta.aplicavel for proposta in propostas)


def test_linhas_vazias_sao_ignoradas() -> None:
    vazia = tuple([None] * len(CABECALHOS))

    assert ler_respostas(CABECALHOS, [vazia, vazia], _catalogo()) == []


def test_le_o_ficheiro_mesmo_com_as_colunas_trocadas() -> None:
    """O fornecedor mexe no ficheiro; o que manda é o título, não a posição."""
    cabecalhos = ["Observações", "Desconto %", "CODIGO", "Preço tabela atualizado"]
    linhas = [("", 20, "PLC0052", 31.2)]

    propostas = ler_respostas(cabecalhos, linhas, _catalogo())

    assert propostas[0].codigo == "PLC0052"
    assert propostas[0].preco_novo == Decimal("31.2")
    assert propostas[0].desconto_novo == Decimal("20")


def test_nao_confunde_preco_atual_com_preco_atualizado() -> None:
    mapa = mapear_colunas(["Preço tabela atual", "Preço tabela atualizado"])

    assert mapa["preco_atual"] == 0
    assert mapa["preco_novo"] == 1


def test_numeros_escritos_a_mao() -> None:
    assert to_decimal("31,20 €") == Decimal("31.20")
    assert to_decimal("1.234,56") == Decimal("1234.56")
    assert to_decimal("  20% ") == Decimal("20")
    assert to_decimal("nao sei") is None


def test_palavras_de_descontinuado() -> None:
    assert diz_descontinuado("Artigo DESCONTINUADO") is True
    assert diz_descontinuado("já não existe") is True
    assert diz_descontinuado("fora de linha") is True
    assert diz_descontinuado("entrega em 5 dias") is False


# ------------------------------------------------------- ciclo completo


@pytest.fixture()
def service(session) -> RespostaFornecedorService:
    return RespostaFornecedorService(session)


def _material_no_catalogo(session):
    return DefMateriaPrimaService(session).criar_materia_prima(
        CriarDefMateriaPrimaData(
            descricao="AGL TERM BEGE ARDENNE 19MM",
            ref_le="PLC0052",
            familia_original_excel="PLACAS",
            unidade="M2",
            preco_tabela=Decimal("30.00"),
            desconto=Decimal("18"),
            preco_liquido=Decimal("24.60"),
            data_ultimo_preco=date(2025, 7, 23),
        )
    )


def test_ciclo_completo_pedido_resposta_catalogo(session, service, tmp_path) -> None:
    """O anexo que sai, preenchido pelo fornecedor, tem de voltar a entrar."""
    from app.domain.pedido_precos import agrupar_por_fornecedor
    from app.services.pedido_precos_service import escrever_anexo

    materia = _material_no_catalogo(session)
    pedido = agrupar_por_fornecedor(
        [SimpleNamespace(**{**materia.__dict__, "fornecedor_id": 1, "ativo": True})],
        [SimpleNamespace(id=1, nome="SONAE", email="c@sonae.pt", email_cc=None, pessoa_contacto=None)],
        HOJE,
    )[0]
    caminho = escrever_anexo(pedido, tmp_path / "pedido.xlsx")

    # O fornecedor preenche o preço e o desconto.
    livro = openpyxl.load_workbook(caminho)
    folha = livro.active
    colunas = {celula.value: celula.column for celula in folha[1]}
    folha.cell(row=2, column=colunas["Preço tabela atualizado"], value=33.00)
    folha.cell(row=2, column=colunas["Desconto %"], value=20)
    livro.save(caminho)

    propostas = service.ler_ficheiro(caminho)
    assert [p.estado for p in propostas] == [ESTADO_ATUALIZA]

    resultado = service.aplicar(propostas, hoje=HOJE)
    assert resultado.atualizadas == 1

    atualizada = DefMateriaPrimaService(session).obter_por_id(materia.id)
    assert atualizada.preco_tabela == Decimal("33.00")
    assert atualizada.desconto == Decimal("20")
    # O líquido é recalculado por nós: 33 × (1 − 20%) = 26,40.
    assert atualizada.preco_liquido == Decimal("26.40")
    assert atualizada.data_ultimo_preco == HOJE

    historico = DefMateriaPrimaService(session).historico_precos(materia.id)
    assert historico[0].origem == "FORNECEDOR"
    assert historico[0].preco_tabela == Decimal("33.00")


def test_aplicar_so_mexe_no_que_vai_na_lista(session, service) -> None:
    materia = _material_no_catalogo(session)
    catalogo = {"PLC0052": SimpleNamespace(id=materia.id, ref_le="PLC0052",
                                           descricao=materia.descricao,
                                           preco_tabela=materia.preco_tabela,
                                           desconto=materia.desconto, margem=None,
                                           referencia_fornecedor=None)}
    propostas = ler_respostas(CABECALHOS, [_linha(preco_novo=31.2)], catalogo)

    resultado = service.aplicar([], hoje=HOJE)
    assert resultado.atualizadas == 0
    assert DefMateriaPrimaService(session).obter_por_id(materia.id).preco_tabela == Decimal("30.00")

    assert service.aplicar(propostas, hoje=HOJE).atualizadas == 1


def test_descontinuado_desativa_sem_apagar(session, service) -> None:
    materia = _material_no_catalogo(session)
    catalogo = {"PLC0052": SimpleNamespace(id=materia.id, ref_le="PLC0052",
                                           descricao=materia.descricao,
                                           preco_tabela=materia.preco_tabela,
                                           desconto=None, margem=None,
                                           referencia_fornecedor=None)}
    propostas = ler_respostas(
        CABECALHOS, [_linha(observacoes="descontinuado")], catalogo
    )

    resultado = service.aplicar(propostas, hoje=HOJE)

    assert resultado.desativadas == 1
    guardada = DefMateriaPrimaService(session).obter_por_id(materia.id)
    assert guardada.ativo is False
    # Continua lá, com o preço que tinha.
    assert guardada.preco_tabela == Decimal("30.00")


def test_ler_folha_encontra_o_cabecalho_abaixo_do_logotipo(tmp_path) -> None:
    livro = openpyxl.Workbook()
    folha = livro.active
    folha["A1"] = "FORNECEDOR XPTO, LDA"
    folha["A2"] = ""
    for coluna, titulo in enumerate(CABECALHOS, start=1):
        folha.cell(row=3, column=coluna, value=titulo)
    folha.cell(row=4, column=1, value="PLC0052")
    caminho = tmp_path / "resposta.xlsx"
    livro.save(caminho)

    cabecalhos, linhas, primeira = ler_folha(caminho)

    assert cabecalhos[0] == "Código"
    assert primeira == 4
    assert linhas[0][0] == "PLC0052"


# ---------------------------------------------------------------- diálogo


def test_dialogo_marca_o_seguro_e_deixa_o_duvidoso_por_marcar() -> None:
    propostas = ler_respostas(
        CABECALHOS,
        [_linha(preco_novo=31.2), _linha(codigo="PLC0053", preco_novo=48.9)],
        {**_catalogo(), **_catalogo(id=2, ref_le="PLC0053")},
    )

    dialogo = RespostaFornecedorDialog(propostas)

    assert dialogo.table.item(0, 0).checkState() == Qt.CheckState.Checked
    assert dialogo.table.item(1, 0).checkState() == Qt.CheckState.Unchecked
    assert len(dialogo.propostas_marcadas()) == 1
    assert "1 com variação invulgar" in dialogo.status_label.text()


def test_dialogo_esconde_o_que_nao_muda_nada() -> None:
    propostas = ler_respostas(
        CABECALHOS,
        [_linha(preco_novo=31.2), _linha(preco_novo=30.00), _linha()],
        _catalogo(),
    )

    dialogo = RespostaFornecedorDialog(propostas)
    assert dialogo.table.rowCount() == 1

    dialogo.mostrar_tudo_input.setChecked(True)
    assert dialogo.table.rowCount() == 3


def test_dialogo_nao_aplica_linhas_que_nao_sao_aplicaveis() -> None:
    propostas = ler_respostas(
        CABECALHOS, [_linha(codigo="XPTO999", preco_novo=10)], _catalogo()
    )
    aplicadas: list = []
    dialogo = RespostaFornecedorDialog(
        propostas, on_aplicar=lambda p: aplicadas.append(p) or True
    )
    dialogo.mostrar_tudo_input.setChecked(True)

    dialogo._aplicar()

    assert aplicadas == []
    assert "nada marcado" in dialogo.status_label.text()
