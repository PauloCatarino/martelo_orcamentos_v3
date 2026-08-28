from __future__ import annotations

from openpyxl import Workbook

from app.models.lista_material_assistente import (
    ListaMaterialPdfDocumento,
    ListaMaterialPdfPreset,
)
from app.services.lista_material_pdf_service import (
    DEFAULT_DOCUMENTS,
    PdfPresetService,
    _export_sheets_to_pdf,
    _remover_ficheiro_a_substituir,
    agrupar_para_exportacao,
    collision_free_path,
    document_filename,
    inspect_pdf_documents,
    normalize_pdf_identifiers,
    read_nome_enc_imos_ix,
    resolve_output_path,
    sync_pdf_document_registry,
)


def test_presets_sao_listados_apenas_para_o_proprio_utilizador(session) -> None:
    service = PdfPresetService(session)
    service.save(
        user_id=10,
        client="JF_VIVA",
        name="Preset Paulo",
        identifiers=["lista_ferragens"],
        export_separate=True,
        create_package=False,
    )
    service.save(
        user_id=20,
        client="JF_VIVA",
        name="Preset Outro Utilizador",
        identifiers=["relatorio"],
        export_separate=True,
        create_package=False,
    )

    assert [row.nome for row in service.list(user_id=10, client="JF_VIVA")] == [
        "Preset Paulo"
    ]


def test_preset_predefinido_e_unico_por_utilizador_e_cliente(session) -> None:
    service = PdfPresetService(session)
    first = service.save(
        user_id=10,
        client="JF_VIVA",
        name="Produção",
        identifiers=["lista_ferragens"],
        export_separate=True,
        create_package=False,
        make_default=True,
    )
    service.save(
        user_id=10,
        client="JF_VIVA",
        name="Arquivo",
        identifiers=["relatorio"],
        export_separate=False,
        create_package=True,
        make_default=False,
    )
    session.refresh(first)
    assert first.predefinido is True

    second = service.save(
        user_id=10,
        client="JF_VIVA",
        name="Arquivo",
        identifiers=["relatorio"],
        export_separate=False,
        create_package=True,
        make_default=True,
    )

    defaults = (
        session.query(ListaMaterialPdfPreset)
        .filter_by(user_id=10, cliente_chave="JF_VIVA", predefinido=True)
        .all()
    )
    assert defaults == [second]
    assert service.list(user_id=10, client="JF_VIVA")[0] == second


def test_exportacao_agrupada_seleciona_ferragens_purch_e_spp(tmp_path) -> None:
    events: list[tuple[object, ...]] = []

    class _Sheet:
        def __init__(self, name: str) -> None:
            self.name = name

        def Select(self, replace: bool) -> None:
            events.append(("select", self.name, replace))

    class _Worksheets:
        def Item(self, name: str) -> _Sheet:
            return _Sheet(name)

    class _Workbook:
        Worksheets = _Worksheets()

    class _ActiveSheet:
        def ExportAsFixedFormat(self, kind: int, output: str) -> None:
            events.append(("export", kind, output))

    class _Excel:
        ActiveSheet = _ActiveSheet()

    output = tmp_path / "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    _export_sheets_to_pdf(
        _Excel(),
        _Workbook(),
        ("1_FERRAGENS", "2_PURCH", "3_SPP"),
        output,
    )

    assert events == [
        ("select", "1_FERRAGENS", True),
        ("select", "2_PURCH", False),
        ("select", "3_SPP", False),
        ("export", 0, str(output)),
    ]


def test_inventario_pdf_mostra_disponiveis_e_indisponiveis(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    workbook.active.title = "DEFENICOES"
    workbook.active["E3"] = "1449_01_26_JF_VIVA"
    workbook.create_sheet("1_FERRAGENS")["A1"] = "Ferragens"
    workbook.create_sheet("2_PURCH")["A1"] = "Purch"
    workbook.save(path)

    states = {item.document.identifier: item for item in inspect_pdf_documents(path)}

    assert states["lista_ferragens"].available is True
    assert states["lista_ferragens"].export_sheets == ("1_FERRAGENS",)
    assert states["lista_purch"].available is True
    assert states["lista_purch"].export_sheets == ("2_PURCH",)
    assert states["lista_spp"].available is False
    assert "3_SPP" in states["lista_spp"].reason
    assert "ferragens" not in states
    assert "caderno_encargos" not in states
    assert "rosto" not in states
    assert "purch" not in states
    assert "spp" not in states
    assert "listagem_cutrite" not in states


def test_nomes_pdf_incluem_nome_enc_imos_ix(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    workbook.active.title = "DEFENICOES"
    workbook.active["E3"] = "1449_01_26_JF_VIVA"
    workbook.save(path)

    by_id = {document.identifier: document for document in DEFAULT_DOCUMENTS}
    nome_enc = read_nome_enc_imos_ix(path)

    assert nome_enc == "1449_01_26_JF_VIVA"
    assert document_filename(by_id["lista_ferragens"], nome_enc) == (
        "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    )
    assert document_filename(by_id["lista_purch"], nome_enc) == (
        "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    )
    assert document_filename(by_id["lista_spp"], nome_enc) == (
        "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    )
    assert document_filename(by_id["resumo_orlas"], nome_enc) == (
        "4_Resumo_Orlas_1449_01_26_JF_VIVA.pdf"
    )
    assert document_filename(by_id["etiqueta_palete"], nome_enc) == (
        "5_Etiqueta_Palete_1449_01_26_JF_VIVA.pdf"
    )
    assert document_filename(by_id["listagem_artigo"], nome_enc) == (
        "6_Lista_Material_1449_01_26_JF_VIVA.pdf"
    )


def test_preset_antigo_de_ferragens_abre_os_tres_documentos() -> None:
    assert normalize_pdf_identifiers(["ferragens"]) == {
        "lista_ferragens",
        "lista_purch",
        "lista_spp",
    }
    assert normalize_pdf_identifiers(["purch", "spp"]) == {
        "lista_ferragens",
        "lista_purch",
        "lista_spp",
    }
    assert normalize_pdf_identifiers(["lista_purch"]) == {"lista_purch"}
    assert normalize_pdf_identifiers(
        ["caderno_encargos", "listagem_cutrite", "relatorio"]
    ) == {"relatorio"}


def test_ferragens_purch_e_spp_sao_vistos_separados_do_mesmo_pdf() -> None:
    by_id = {document.identifier: document for document in DEFAULT_DOCUMENTS}
    trio = ("lista_ferragens", "lista_purch", "lista_spp")

    assert by_id["lista_ferragens"].sheets == ("1_FERRAGENS",)
    assert by_id["lista_purch"].sheets == ("2_PURCH",)
    assert by_id["lista_spp"].sheets == ("3_SPP",)
    assert by_id["lista_purch"].name == "Purch (Objectos Comprados)"
    assert by_id["lista_spp"].name == "SPP (Stretchable Purchased Part)"
    assert {by_id[identifier].category for identifier in trio} == {"Ferragens"}
    # Vistos separados no menu, um único ficheiro na pasta.
    assert {by_id[identifier].group for identifier in trio} == {"ferragens"}
    assert {by_id[identifier].filename for identifier in trio} == {
        "2_Lista_Ferragens_{nome_enc_imos}.pdf"
    }


def test_ferragens_purch_e_spp_saem_num_unico_pdf(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    workbook.active.title = "DEFENICOES"
    workbook.active["E3"] = "1449_01_26_JF_VIVA"
    for nome in ("1_FERRAGENS", "2_PURCH", "3_SPP", "ResumoOrlas"):
        workbook.create_sheet(nome)["A1"] = nome
    workbook.save(path)

    states = [
        state
        for state in inspect_pdf_documents(path)
        if state.document.identifier
        in ("lista_ferragens", "lista_purch", "lista_spp", "resumo_orlas")
    ]

    assert agrupar_para_exportacao(states, "1449_01_26_JF_VIVA") == [
        (
            "Ferragens + Purch + SPP",
            "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf",
            ("1_FERRAGENS", "2_PURCH", "3_SPP"),
        ),
        (
            "Resumo de Orlas",
            "4_Resumo_Orlas_1449_01_26_JF_VIVA.pdf",
            ("ResumoOrlas",),
        ),
    ]


def test_grupo_de_ferragens_leva_so_os_separadores_escolhidos(tmp_path) -> None:
    path = tmp_path / "lista.xlsx"
    workbook = Workbook()
    workbook.active.title = "DEFENICOES"
    workbook.active["E3"] = "1449_01_26_JF_VIVA"
    for nome in ("1_FERRAGENS", "2_PURCH", "3_SPP"):
        workbook.create_sheet(nome)["A1"] = nome
    workbook.save(path)

    states = {state.document.identifier: state for state in inspect_pdf_documents(path)}
    escolhidos = [states["lista_ferragens"], states["lista_spp"]]

    assert agrupar_para_exportacao(escolhidos, "1449_01_26_JF_VIVA") == [
        (
            "Ferragens + Purch + SPP",
            "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf",
            ("1_FERRAGENS", "3_SPP"),
        )
    ]


def test_substituir_devolve_o_nome_original(tmp_path) -> None:
    existing = tmp_path / "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    existing.write_bytes(b"existente")

    substituir = resolve_output_path(
        tmp_path, "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf", overwrite=True
    )
    manter = resolve_output_path(
        tmp_path, "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf", overwrite=False
    )

    assert substituir == existing
    assert manter.name == "2_Lista_Ferragens_1449_01_26_JF_VIVA_2.pdf"


def test_substituir_apaga_o_pdf_antigo_antes_de_exportar(tmp_path) -> None:
    existing = tmp_path / "2_Lista_Purch_1449_01_26_JF_VIVA.pdf"
    existing.write_bytes(b"existente")

    _remover_ficheiro_a_substituir(existing, True)
    assert existing.exists() is False

    mantido = tmp_path / "2_Lista_SPP_1449_01_26_JF_VIVA.pdf"
    mantido.write_bytes(b"existente")
    _remover_ficheiro_a_substituir(mantido, False)
    assert mantido.read_bytes() == b"existente"


def test_colisao_de_nome_nunca_sobrescreve(tmp_path) -> None:
    existing = tmp_path / "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    existing.write_bytes(b"existente")

    result = collision_free_path(
        tmp_path, "2_Lista_Ferragens_1449_01_26_JF_VIVA.pdf"
    )

    assert result.name == "2_Lista_Ferragens_1449_01_26_JF_VIVA_2.pdf"
    assert existing.read_bytes() == b"existente"


def test_registo_pdf_e_sincronizado_sem_apagar_personalizados(session) -> None:
    session.add(
        ListaMaterialPdfDocumento(
            identificador="personalizado",
            nome="Meu documento",
            categoria="Produção",
            origem_tipo="folha",
            origem_valor="MINHA_FOLHA",
            nome_ficheiro="Meu.pdf",
        )
    )
    session.commit()

    assert sync_pdf_document_registry(session) == len(DEFAULT_DOCUMENTS)
    assert session.query(ListaMaterialPdfDocumento).count() == len(DEFAULT_DOCUMENTS) + 1
    assert session.query(ListaMaterialPdfDocumento).filter_by(identificador="personalizado").one()


def test_registo_desativa_documentos_retirados_do_centro(session) -> None:
    session.add(
        ListaMaterialPdfDocumento(
            identificador="listagem_cutrite",
            nome="Listagem CUT-RITE",
            categoria="CUT-RITE",
            origem_tipo="folha",
            origem_valor="LISTAGEM_CUT_RITE",
            nome_ficheiro="Listagem_CUT_RITE.pdf",
            ativo=True,
        )
    )
    session.commit()

    sync_pdf_document_registry(session)

    retired = session.query(ListaMaterialPdfDocumento).filter_by(
        identificador="listagem_cutrite"
    ).one()
    assert retired.ativo is False
